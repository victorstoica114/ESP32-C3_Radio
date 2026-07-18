from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def read_session(result_dir: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    metadata = json.loads((result_dir / "metadata.json").read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    with (result_dir / "summary.csv").open(encoding="utf-8", newline="") as stream:
        for raw in csv.DictReader(stream):
            row: dict[str, Any] = dict(raw)
            for field in (
                "tx_power_dbm",
                "bit_rate_kbps",
                "requested_duration_s",
                "actual_tx_duration_ms",
                "frames_transmitted",
                "bytes_transmitted",
                "captured_samples",
                "active_samples",
                "active_window_s",
                "sample_loss_percent",
                "baseline_mean_uA",
                "baseline_median_uA",
                "mean_current_uA",
                "stdev_current_uA",
                "peak_current_uA",
                "mean_excess_current_uA",
                "mean_power_mW",
                "stdev_power_mW",
                "peak_power_mW",
                "mean_excess_power_mW",
                "energy_60s_mJ",
            ):
                row[field] = float(raw[field])
            row["frames_received"] = (
                int(raw["frames_received"]) if raw["frames_received"] else ""
            )
            row["frame_loss_percent"] = (
                float(raw["frame_loss_percent"])
                if raw["frame_loss_percent"]
                else ""
            )
            rows.append(row)
    return rows, metadata


def validate(
    tx_rows: list[dict[str, Any]],
    tx_metadata: dict[str, Any],
    rx_rows: list[dict[str, Any]],
    rx_metadata: dict[str, Any],
) -> None:
    if tx_metadata.get("measurement_direction") != "tx":
        raise ValueError("First continuous session is not TX")
    if rx_metadata.get("measurement_direction") != "rx":
        raise ValueError("Second continuous session is not RX")
    tx_powers = tuple(sorted(row["tx_power_dbm"] for row in tx_rows))
    rx_powers = tuple(sorted(row["tx_power_dbm"] for row in rx_rows))
    if len(set(tx_powers)) != len(tx_powers) or tx_powers != rx_powers:
        raise ValueError(
            f"TX/RX power sweeps are incomplete or different: "
            f"TX={tx_powers}, RX={rx_powers}"
        )
    for rows, direction in ((tx_rows, "tx"), (rx_rows, "rx")):
        if any(row["status"] != "ok" for row in rows):
            raise ValueError(f"{direction.upper()} contains non-OK rows")
        if any(row["requested_duration_s"] != 60.0 for row in rows):
            raise ValueError(f"{direction.upper()} rows are not 60-second measurements")
    shared_fields = (
        "bit_rate_kbps",
        "frame_bytes",
        "content_bytes_per_frame",
        "inter_frame_gap_ms",
        "voltage_mv",
    )
    for field in shared_fields:
        if {row[field] for row in tx_rows} != {row[field] for row in rx_rows}:
            raise ValueError(f"TX/RX mismatch in {field}")
    if any(row.get("rf_profile") for row in tx_rows + rx_rows):
        if {row.get("rf_profile", "") for row in tx_rows} != {
            row.get("rf_profile", "") for row in rx_rows
        }:
            raise ValueError("TX/RX mismatch in rf_profile")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = list(rows[0])
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(42, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def write_xlsx(
    path: Path,
    tx_rows: list[dict[str, Any]],
    rx_rows: list[dict[str, Any]],
    tx_metadata: dict[str, Any],
    rx_metadata: dict[str, Any],
) -> None:
    workbook = Workbook()
    results = workbook.active
    results.title = "continuous_results"
    fields = list(tx_rows[0])
    results.append(fields)
    for row in tx_rows + rx_rows:
        results.append([row.get(field, "") for field in fields])
    _style_sheet(results)

    comparison = workbook.create_sheet("comparison")
    comparison_fields = [
        "tx_power_dbm",
        "tx_mean_current_mA",
        "tx_mean_power_mW",
        "tx_excess_power_mW",
        "tx_energy_60s_J",
        "rx_mean_current_mA",
        "rx_mean_power_mW",
        "rx_excess_power_mW",
        "rx_energy_60s_J",
        "rx_frames_received",
        "rx_frames_transmitted",
        "rx_frame_loss_percent",
    ]
    comparison.append(comparison_fields)
    for power in sorted(row["tx_power_dbm"] for row in tx_rows):
        tx = next(row for row in tx_rows if row["tx_power_dbm"] == power)
        rx = next(row for row in rx_rows if row["tx_power_dbm"] == power)
        comparison.append(
            [
                power,
                tx["mean_current_uA"] / 1000.0,
                tx["mean_power_mW"],
                tx["mean_excess_power_mW"],
                tx["energy_60s_mJ"] / 1000.0,
                rx["mean_current_uA"] / 1000.0,
                rx["mean_power_mW"],
                rx["mean_excess_power_mW"],
                rx["energy_60s_mJ"] / 1000.0,
                rx["frames_received"],
                int(rx["frames_transmitted"]),
                rx["frame_loss_percent"],
            ]
        )
    _style_sheet(comparison)

    metadata_sheet = workbook.create_sheet("metadata")
    metadata_sheet.append(["field", "TX", "RX"])
    for field in (
        "created_utc",
        "profile_id",
        "module",
        "measurement_direction",
        "measured_port",
        "peer_port",
        "ppk_port",
        "voltage_mv",
        "sample_rate_hz",
        "bit_rate_kbps",
        "frame_bytes",
        "content_bytes_per_frame",
        "inter_frame_gap_ms",
        "requested_duration_s",
    ):
        metadata_sheet.append([field, tx_metadata.get(field, ""), rx_metadata.get(field, "")])
    _style_sheet(metadata_sheet)
    workbook.save(path)


def _number(value: float | int) -> str:
    return f"{value:.9g}"


def _coordinates(rows: list[dict[str, Any]], field: str) -> str:
    return " ".join(
        f"({_number(row['tx_power_dbm'])},{_number(float(row[field]))})"
        for row in sorted(rows, key=lambda item: item["tx_power_dbm"])
    )


def _rate_description(row: dict[str, Any]) -> str:
    rate = _number(float(row["bit_rate_kbps"])).replace(".", "{,}")
    profile = str(row.get("rf_profile") or "")
    return f"{profile} ({rate} kbps)" if profile else f"{rate} kbps"


def write_power_tex(
    path: Path,
    tx_rows: list[dict[str, Any]],
    rx_rows: list[dict[str, Any]],
    module_title: str,
) -> None:
    powers = sorted(row["tx_power_dbm"] for row in tx_rows)
    ticks = ",".join(_number(power) for power in powers)
    xmin = min(powers) - max(2.0, (max(powers) - min(powers)) * 0.10)
    xmax = max(powers) + max(2.0, (max(powers) - min(powers)) * 0.10)
    rate_label = _rate_description(tx_rows[0])
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\usepgfplotslibrary{groupplots}",
        r"\usetikzlibrary{calc}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{groupplot}[/pgf/number format/use comma,",
        r"  group style={group size=2 by 1, horizontal sep=1.7cm},",
        r"  width=8.2cm, height=6.4cm,",
        rf"  xmin={_number(xmin)}, xmax={_number(xmax)}, xtick={{{ticks}}},",
        r"  xlabel={Puterea RF configurată [dBm]},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"  every axis plot/.append style={line width=1.2pt, mark size=2.6pt},",
        r"  legend style={font=\small, draw=none, fill=none,",
        r"    at={(0.5,-0.25)}, anchor=north, legend columns=2},",
        r"]",
        r"\nextgroupplot[title={Putere totală}, ylabel={Puterea electrică medie [mW]}]",
        rf"\addplot+[blue!75!black, mark=*] coordinates {{{_coordinates(tx_rows, 'mean_power_mW')}}};",
        r"\addlegendentry{TX}",
        rf"\addplot+[red!75!black, dashed, mark=square*] coordinates {{{_coordinates(rx_rows, 'mean_power_mW')}}};",
        r"\addlegendentry{RX}",
        r"\nextgroupplot[title={Putere peste standby}, ylabel={Puterea medie suplimentară [mW]}]",
        rf"\addplot+[blue!75!black, mark=*] coordinates {{{_coordinates(tx_rows, 'mean_excess_power_mW')}}};",
        r"\addlegendentry{TX}",
        rf"\addplot+[red!75!black, dashed, mark=square*] coordinates {{{_coordinates(rx_rows, 'mean_excess_power_mW')}}};",
        r"\addlegendentry{RX}",
        r"\end{groupplot}",
        r"\node[font=\bfseries\large] at ($(group c1r1.north)!0.5!(group c2r1.north)+(0,0.85cm)$)",
        rf"  {{{module_title}: putere medie în flux continuu}};",
        r"\node[font=\footnotesize, align=center, text width=16.5cm]",
        r"  at ($(group c1r1.south)!0.5!(group c2r1.south)+(0,-2.55cm)$)",
        rf"  {{Fiecare punct este media unei ferestre de 60 s la 3,3 V; profil {rate_label},\\",
        r"   pauză de 15 ms între cadre. Puterea RX de pe axa X este puterea emițătorului de stimul.};",
        r"\end{tikzpicture}",
        r"\end{document}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def write_delivery_tex(
    path: Path, rx_rows: list[dict[str, Any]], module_title: str
) -> None:
    delivery_rows = [
        {**row, "delivery_percent": 100.0 - float(row["frame_loss_percent"])}
        for row in rx_rows
    ]
    powers = sorted(row["tx_power_dbm"] for row in rx_rows)
    ticks = ",".join(_number(power) for power in powers)
    xmin = min(powers) - max(2.0, (max(powers) - min(powers)) * 0.10)
    xmax = max(powers) + max(2.0, (max(powers) - min(powers)) * 0.10)
    rate_label = _rate_description(rx_rows[0])
    frame_counts = [int(row["frames_transmitted"]) for row in rx_rows]
    frame_count_note = (
        str(frame_counts[0])
        if min(frame_counts) == max(frame_counts)
        else f"{min(frame_counts)}--{max(frame_counts)}"
    )
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{axis}[width=10.5cm, height=6.4cm,",
        rf"  title={{{module_title}: cadre recepționate în testul continuu}},",
        rf"  xmin={_number(xmin)}, xmax={_number(xmax)}, xtick={{{ticks}}},",
        r"  ymin=0, ymax=105, ytick={0,20,40,60,80,100},",
        r"  xlabel={Puterea emițătorului de stimul [dBm]},",
        r"  ylabel={Cadre recepționate [\%]},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"]",
        rf"\addplot+[red!75!black, line width=1.2pt, mark=square*, mark size=2.8pt] coordinates {{{_coordinates(delivery_rows, 'delivery_percent')}}};",
        r"\end{axis}",
        r"\node[font=\footnotesize, align=center, text width=10cm] at (5.25,-1.45)",
        rf"  {{{frame_count_note} cadre transmise/punct; {int(rx_rows[0]['frame_bytes'])} B/cadru, profil {rate_label}, 60 s.}};",
        r"\end{tikzpicture}",
        r"\end{document}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate continuous TX/RX reports and PGFPlots figures")
    parser.add_argument("tx_result_dir", type=Path)
    parser.add_argument("rx_result_dir", type=Path)
    parser.add_argument("output_base", type=Path)
    args = parser.parse_args()

    tx_rows, tx_metadata = read_session(args.tx_result_dir)
    rx_rows, rx_metadata = read_session(args.rx_result_dir)
    validate(tx_rows, tx_metadata, rx_rows, rx_metadata)
    module_title = str(tx_metadata["module"])
    if rx_metadata.get("module") != module_title:
        raise ValueError("TX and RX continuous sessions refer to different modules")
    args.output_base.parent.mkdir(parents=True, exist_ok=True)
    combined = sorted(
        tx_rows + rx_rows,
        key=lambda row: (row["measurement_direction"], row["tx_power_dbm"]),
    )
    csv_path = args.output_base.with_suffix(".csv")
    xlsx_path = args.output_base.with_suffix(".xlsx")
    power_tex = args.output_base.with_name(args.output_base.name + "_average_power").with_suffix(".tex")
    delivery_tex = args.output_base.with_name(args.output_base.name + "_rx_delivery").with_suffix(".tex")
    write_csv(csv_path, combined)
    write_xlsx(xlsx_path, tx_rows, rx_rows, tx_metadata, rx_metadata)
    write_power_tex(power_tex, tx_rows, rx_rows, module_title)
    write_delivery_tex(delivery_tex, rx_rows, module_title)
    print(csv_path.resolve())
    print(xlsx_path.resolve())
    print(power_tex.resolve())
    print(delivery_tex.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
