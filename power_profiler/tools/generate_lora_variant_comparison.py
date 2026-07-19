from __future__ import annotations

import argparse
import csv
import json
import statistics
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from render_lora_campaign_plots import (
    BLUE,
    GREEN,
    ORANGE,
    RED,
    PdfCanvas,
    _axes,
    _legend,
    _plot_series,
)


CONTINUOUS_FIELDS = (
    "measurement_direction",
    "tx_power_dbm",
    "spreading_factor",
    "bandwidth_khz",
    "classic_mean_current_mA",
    "variant_mean_current_mA",
    "current_delta_percent",
    "classic_mean_power_mW",
    "variant_mean_power_mW",
    "power_delta_percent",
    "classic_frames_transmitted",
    "classic_frames_received",
    "variant_frames_transmitted",
    "variant_frames_received",
)

PACKET_FIELDS = (
    "measurement_direction",
    "payload_bytes",
    "tx_power_dbm",
    "spreading_factor",
    "bandwidth_khz",
    "classic_energy_mJ",
    "variant_energy_mJ",
    "energy_delta_percent",
    "classic_energy_cv_percent",
    "variant_energy_cv_percent",
    "classic_packets_received",
    "variant_packets_received",
)


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        return list(csv.DictReader(stream))


def _find_report(directory: Path, suffix: str) -> Path:
    matches = sorted(directory.glob(f"*{suffix}"))
    if len(matches) != 1:
        raise ValueError(
            f"Expected one *{suffix} report in {directory}, found {len(matches)}"
        )
    return matches[0]


def _f(value: Any) -> float:
    return float(value or 0)


def _delta(reference: float, variant: float) -> float:
    return (variant / reference - 1.0) * 100.0 if reference else 0.0


def _continuous_key(row: dict[str, str]) -> tuple[str, float, int, float]:
    return (
        row["measurement_direction"],
        _f(row["tx_power_dbm"]),
        int(_f(row["spreading_factor"])),
        _f(row["bandwidth_khz"]),
    )


def _packet_key(row: dict[str, str]) -> tuple[str, int, float, int, float]:
    return (
        row["measurement_direction"],
        int(_f(row["payload_bytes"])),
        _f(row["tx_power_dbm"]),
        int(_f(row["spreading_factor"])),
        _f(row["bandwidth_khz"]),
    )


def compare_continuous(
    classic: Iterable[dict[str, str]], variant: Iterable[dict[str, str]]
) -> list[dict[str, Any]]:
    classic_by_key = {_continuous_key(row): row for row in classic}
    variant_by_key = {_continuous_key(row): row for row in variant}
    if classic_by_key.keys() != variant_by_key.keys():
        raise ValueError("Continuous reports do not describe identical test points")
    result = []
    for key in sorted(classic_by_key):
        direction, power, spreading_factor, bandwidth = key
        left, right = classic_by_key[key], variant_by_key[key]
        left_current = _f(left["mean_current_uA"]) / 1000.0
        right_current = _f(right["mean_current_uA"]) / 1000.0
        left_power = _f(left["mean_power_mW"])
        right_power = _f(right["mean_power_mW"])
        result.append(
            {
                "measurement_direction": direction,
                "tx_power_dbm": power,
                "spreading_factor": spreading_factor,
                "bandwidth_khz": bandwidth,
                "classic_mean_current_mA": left_current,
                "variant_mean_current_mA": right_current,
                "current_delta_percent": _delta(left_current, right_current),
                "classic_mean_power_mW": left_power,
                "variant_mean_power_mW": right_power,
                "power_delta_percent": _delta(left_power, right_power),
                "classic_frames_transmitted": int(_f(left["frames_transmitted"])),
                "classic_frames_received": int(_f(left["frames_received"])),
                "variant_frames_transmitted": int(_f(right["frames_transmitted"])),
                "variant_frames_received": int(_f(right["frames_received"])),
            }
        )
    return result


def compare_packets(
    classic: Iterable[dict[str, str]], variant: Iterable[dict[str, str]]
) -> list[dict[str, Any]]:
    classic_by_key = {_packet_key(row): row for row in classic}
    variant_by_key = {_packet_key(row): row for row in variant}
    if classic_by_key.keys() != variant_by_key.keys():
        raise ValueError("Packet reports do not describe identical test points")
    result = []
    for key in sorted(classic_by_key):
        direction, payload, power, spreading_factor, bandwidth = key
        left, right = classic_by_key[key], variant_by_key[key]
        left_energy = _f(left["energy_total_mJ_mean"])
        right_energy = _f(right["energy_total_mJ_mean"])
        result.append(
            {
                "measurement_direction": direction,
                "payload_bytes": payload,
                "tx_power_dbm": power,
                "spreading_factor": spreading_factor,
                "bandwidth_khz": bandwidth,
                "classic_energy_mJ": left_energy,
                "variant_energy_mJ": right_energy,
                "energy_delta_percent": _delta(left_energy, right_energy),
                "classic_energy_cv_percent": _f(left["energy_cv_percent"]),
                "variant_energy_cv_percent": _f(right["energy_cv_percent"]),
                "classic_packets_received": int(_f(left["packets_received"])),
                "variant_packets_received": int(_f(right["packets_received"])),
            }
        )
    return result


def _write_csv(path: Path, fields: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(
    path: Path,
    continuous: list[dict[str, Any]],
    packets: list[dict[str, Any]],
    summary: dict[str, Any],
) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, fields, rows in (
        ("summary", tuple(summary), [summary]),
        ("continuous", CONTINUOUS_FIELDS, continuous),
        ("packet_energy", PACKET_FIELDS, packets),
    ):
        sheet = workbook.create_sheet(title)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(rows) + 1}"
        for column, field in enumerate(fields, 1):
            cell = sheet.cell(1, column, field)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for row_index, row in enumerate(rows, 2):
            for column, field in enumerate(fields, 1):
                sheet.cell(row_index, column, row.get(field))
        for column, field in enumerate(fields, 1):
            width = max(len(field), *(len(str(row.get(field, ""))) for row in rows))
            sheet.column_dimensions[get_column_letter(column)].width = min(width + 2, 36)
    workbook.save(path)


def _averaged_rx(rows: list[dict[str, Any]], field: str) -> list[tuple[float, float]]:
    powers = sorted({row["tx_power_dbm"] for row in rows})
    return [
        (
            power,
            statistics.mean(
                row[field]
                for row in rows
                if row["measurement_direction"] == "rx"
                and row["tx_power_dbm"] == power
            ),
        )
        for power in powers
    ]


def _current_plot(path: Path, rows: list[dict[str, Any]]) -> None:
    canvas = PdfCanvas(1080, 540)
    canvas.text(315, 505, "RA-02 CLASSIC VS RA-02 + 2CAP", size=16, bold=True)
    boxes = ((85, 95, 420, 340), (610, 95, 385, 340))
    powers = sorted({row["tx_power_dbm"] for row in rows})
    ticks = tuple((power, f"{power:g}") for power in powers)
    tx = [row for row in rows if row["measurement_direction"] == "tx"]
    _axes(
        canvas,
        boxes[0],
        y_ticks=(0, 10, 20, 30, 40, 50, 60, 70),
        y_range=(0, 70),
        x_ticks=ticks,
        x_range=(min(powers), max(powers)),
        y_label="TX current [mA]",
    )
    _plot_series(
        canvas,
        boxes[0],
        [(row["tx_power_dbm"], row["classic_mean_current_mA"]) for row in tx],
        x_range=(min(powers), max(powers)),
        y_range=(0, 70),
        color=BLUE,
        marker=0,
    )
    _plot_series(
        canvas,
        boxes[0],
        [(row["tx_power_dbm"], row["variant_mean_current_mA"]) for row in tx],
        x_range=(min(powers), max(powers)),
        y_range=(0, 70),
        color=RED,
        marker=1,
        dash="[7 3] 0",
    )
    _axes(
        canvas,
        boxes[1],
        y_ticks=(11, 11.5, 12, 12.5, 13),
        y_range=(11, 13),
        x_ticks=ticks,
        x_range=(min(powers), max(powers)),
        y_label="RX current [mA]",
    )
    _plot_series(
        canvas,
        boxes[1],
        _averaged_rx(rows, "classic_mean_current_mA"),
        x_range=(min(powers), max(powers)),
        y_range=(11, 13),
        color=BLUE,
        marker=0,
    )
    _plot_series(
        canvas,
        boxes[1],
        _averaged_rx(rows, "variant_mean_current_mA"),
        x_range=(min(powers), max(powers)),
        y_range=(11, 13),
        color=RED,
        marker=1,
        dash="[7 3] 0",
    )
    _legend(canvas, (("Classic", BLUE, 0), ("2Cap", RED, 1)), 435, 55)
    canvas.text(445, 25, "Transmitter power [dBm]", size=9)
    path.parent.mkdir(parents=True, exist_ok=True)
    canvas.save(path)


def _energy_delta_plot(path: Path, rows: list[dict[str, Any]]) -> None:
    tx = [row for row in rows if row["measurement_direction"] == "tx"]
    powers = sorted({row["tx_power_dbm"] for row in tx})
    spreading_factors = sorted({row["spreading_factor"] for row in tx})
    colors = (BLUE, ORANGE, GREEN)
    canvas = PdfCanvas(1080, 540)
    canvas.text(280, 505, "RA-02 + 2CAP TX ENERGY DELTA VS CLASSIC", size=16, bold=True)
    boxes = [(75 + index * 345, 95, 285, 340) for index in range(3)]
    for box, power in zip(boxes, powers):
        _axes(
            canvas,
            box,
            y_ticks=(-5, -4, -3, -2, -1, 0, 1),
            y_range=(-5, 1),
            x_ticks=((8, "8"), (32, "32"), (128, "128")),
            x_range=(8, 128),
            log_x=True,
            y_label="Energy delta [%]",
        )
        canvas.text(box[0] + 115, 455, f"{power:g} dBm", size=11, bold=True)
        for index, spreading_factor in enumerate(spreading_factors):
            selected = sorted(
                (
                    row
                    for row in tx
                    if row["tx_power_dbm"] == power
                    and row["spreading_factor"] == spreading_factor
                ),
                key=lambda row: row["payload_bytes"],
            )
            _plot_series(
                canvas,
                box,
                [(row["payload_bytes"], row["energy_delta_percent"]) for row in selected],
                x_range=(8, 128),
                y_range=(-5, 1),
                color=colors[index],
                marker=index,
                log_x=True,
                dash=("[] 0", "[7 3] 0", "[8 3 2 3] 0")[index],
            )
    _legend(
        canvas,
        [(f"SF{sf}", colors[index], index) for index, sf in enumerate(spreading_factors)],
        420,
        55,
    )
    canvas.text(420, 25, "Payload [B]; negative means lower 2Cap energy", size=9)
    canvas.save(path)


def _write_current_tex(path: Path, rows: list[dict[str, Any]]) -> None:
    tx = [row for row in rows if row["measurement_direction"] == "tx"]
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage{pgfplots}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{axis}[width=0.92\linewidth,height=7cm,grid=both,",
        r"xlabel={TX power (dBm)},ylabel={Mean TX current (mA)},",
        r"legend pos=north west,title={RA-02 classic vs. RA-02 + 2Cap}]",
    ]
    for label, field, style in (
        ("RA-02 classic", "classic_mean_current_mA", "blue,mark=*"),
        ("RA-02 + 2Cap", "variant_mean_current_mA", "red,dashed,mark=square*"),
    ):
        coordinates = " ".join(
            f"({row['tx_power_dbm']:g},{row[field]:.6f})" for row in tx
        )
        lines.extend((rf"\addplot+[{style}] coordinates {{{coordinates}}};", rf"\addlegendentry{{{label}}}"))
    lines.extend((r"\end{axis}", r"\end{tikzpicture}", r"\end{document}"))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_energy_delta_tex(path: Path, rows: list[dict[str, Any]]) -> None:
    tx = [row for row in rows if row["measurement_direction"] == "tx"]
    powers = sorted({row["tx_power_dbm"] for row in tx})
    spreading_factors = sorted({row["spreading_factor"] for row in tx})
    styles = (
        "blue,solid,mark=*",
        "orange,dashed,mark=square*",
        "green!50!black,dashdotted,mark=triangle*",
    )
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage{pgfplots}",
        r"\usepgfplotslibrary{groupplots}",
        r"\usetikzlibrary{calc}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{groupplot}[group style={group size=3 by 1,horizontal sep=1.25cm},",
        r"width=5.8cm,height=6.2cm,xmode=log,log basis x=2,ymin=-5,ymax=1,",
        r"xtick={8,32,128},xticklabels={8,32,128},grid=both,xlabel={Payload [B]},",
        r"legend style={draw=none,fill=none,font=\small,at={(0.5,-0.27)},anchor=north,legend columns=-1}]",
    ]
    for power_index, power in enumerate(powers):
        ylabel = (
            r",ylabel={2Cap energy delta vs. classic [\%]}"
            if power_index == 0
            else ""
        )
        lines.append(rf"\nextgroupplot[title={{{power:g} dBm}}{ylabel}]")
        for sf_index, spreading_factor in enumerate(spreading_factors):
            selected = sorted(
                (
                    row
                    for row in tx
                    if row["tx_power_dbm"] == power
                    and row["spreading_factor"] == spreading_factor
                ),
                key=lambda row: row["payload_bytes"],
            )
            coordinates = " ".join(
                f"({row['payload_bytes']:g},{row['energy_delta_percent']:.6f})"
                for row in selected
            )
            lines.append(
                rf"\addplot+[{styles[sf_index]}] coordinates {{{coordinates}}};"
            )
            if power_index == 1:
                lines.append(rf"\addlegendentry{{SF{spreading_factor}}}")
    lines.extend(
        (
            r"\end{groupplot}",
            r"\node[font=\bfseries\large] at ($(group c2r1.north)+(0,0.85cm)$) {RA-02 + 2Cap TX energy delta versus classic RA-02};",
            r"\node[font=\footnotesize] at ($(group c2r1.south)+(0,-2.75cm)$) {Negative values mean lower energy for the 2Cap pair.};",
            r"\end{tikzpicture}",
            r"\end{document}",
        )
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generate(
    classic_dir: Path,
    variant_dir: Path,
    output_dir: Path,
    base_name: str,
) -> list[Path]:
    classic_continuous = _read_csv(_find_report(classic_dir, "_continuous.csv"))
    variant_continuous = _read_csv(_find_report(variant_dir, "_continuous.csv"))
    classic_packets = _read_csv(_find_report(classic_dir, "_data.csv"))
    variant_packets = _read_csv(_find_report(variant_dir, "_data.csv"))
    continuous = compare_continuous(classic_continuous, variant_continuous)
    packets = compare_packets(classic_packets, variant_packets)
    tx_packet_deltas = [
        row["energy_delta_percent"]
        for row in packets
        if row["measurement_direction"] == "tx"
    ]
    rx_packet_deltas = [
        row["energy_delta_percent"]
        for row in packets
        if row["measurement_direction"] == "rx"
    ]
    summary = {
        "matching_continuous_points": len(continuous),
        "matching_packet_points": len(packets),
        "tx_packet_points_lower_with_2cap": sum(delta < 0 for delta in tx_packet_deltas),
        "tx_packet_points_total": len(tx_packet_deltas),
        "tx_packet_energy_delta_mean_percent": statistics.mean(tx_packet_deltas),
        "tx_packet_energy_delta_min_percent": min(tx_packet_deltas),
        "tx_packet_energy_delta_max_percent": max(tx_packet_deltas),
        "rx_packet_energy_delta_mean_percent": statistics.mean(rx_packet_deltas),
    }
    output_dir.mkdir(parents=True, exist_ok=True)
    continuous_path = output_dir / f"{base_name}_continuous.csv"
    packet_path = output_dir / f"{base_name}_packet_energy.csv"
    workbook_path = output_dir / f"{base_name}.xlsx"
    summary_path = output_dir / f"{base_name}_summary.json"
    current_plot = output_dir / f"{base_name}_average_current.pdf"
    energy_plot = output_dir / f"{base_name}_tx_energy_delta.pdf"
    tex_path = output_dir / f"{base_name}_average_current.tex"
    energy_tex_path = output_dir / f"{base_name}_tx_energy_delta.tex"
    _write_csv(continuous_path, CONTINUOUS_FIELDS, continuous)
    _write_csv(packet_path, PACKET_FIELDS, packets)
    _write_xlsx(workbook_path, continuous, packets, summary)
    summary_path.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    _current_plot(current_plot, continuous)
    _energy_delta_plot(energy_plot, packets)
    _write_current_tex(tex_path, continuous)
    _write_energy_delta_tex(energy_tex_path, packets)
    return [
        continuous_path,
        packet_path,
        workbook_path,
        summary_path,
        current_plot,
        current_plot.with_suffix(".png"),
        energy_plot,
        energy_plot.with_suffix(".png"),
        tex_path,
        energy_tex_path,
    ]


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare two matching LoRa campaigns")
    parser.add_argument("classic_dir", type=Path)
    parser.add_argument("variant_dir", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--base-name", default="lora_variant_comparison")
    args = parser.parse_args()
    for path in generate(
        args.classic_dir, args.variant_dir, args.output_dir, args.base_name
    ):
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
