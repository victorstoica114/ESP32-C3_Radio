from __future__ import annotations

import argparse
import copy
import csv
import json
import math
import shutil
import statistics
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


COLORS = ("blue!75!black", "orange!90!black", "green!50!black")
STYLES = ("solid", "dashed", "dashdotted")
MARKERS = ("*", "square*", "triangle*")
PACKET_FIELDS = (
    "profile_id",
    "module",
    "measurement_direction",
    "payload_bytes",
    "tx_power_dbm",
    "spreading_factor",
    "bandwidth_khz",
    "runs",
    "events_detected",
    "packets_attempted",
    "packets_received",
    "packets_lost",
    "packet_loss_percent",
    "event_duration_ms_mean",
    "event_duration_ms_stdev",
    "event_mean_uA_mean",
    "event_peak_uA_mean",
    "energy_total_uJ_mean",
    "energy_total_uJ_stdev",
    "energy_total_mJ_mean",
    "energy_cv_percent",
    "sample_loss_percent_mean",
    "sample_loss_percent_max",
    "source_directory",
)
LOSS_FIELDS = (
    "spreading_factor",
    "bandwidth_khz",
    "effective_payload_rate_kbps",
    "tx_power_dbm",
    "frames_transmitted",
    "frames_received",
    "frames_lost",
    "frame_loss_percent",
    "delivery_percent",
    "requested_duration_s",
    "frame_bytes",
    "inter_frame_gap_ms",
    "status",
    "source_directory",
)


def _f(value: Any) -> float:
    return float(value) if value not in (None, "") else 0.0


def _bandwidth_khz(parameters: dict[str, Any]) -> float:
    if parameters.get("bandwidth_khz") not in (None, ""):
        return _f(parameters["bandwidth_khz"])
    if parameters.get("bandwidth_hz") not in (None, ""):
        return _f(parameters["bandwidth_hz"]) / 1000.0
    raise ValueError("LoRa result does not contain a bandwidth parameter")


def _n(value: float | int) -> str:
    return f"{value:.9g}"


def _latex_n(value: float | int) -> str:
    return _n(value).replace(".", "{,}")


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def _manifest(path: Path) -> dict[str, Any]:
    value = copy.deepcopy(json.loads(path.read_text(encoding="utf-8")))
    if value.get("kind") != "campaign":
        raise ValueError("Expected a campaign manifest")

    recovery_path = path.parent / "recovery_overrides.json"
    if recovery_path.exists():
        recovery = json.loads(recovery_path.read_text(encoding="utf-8"))
        overrides = recovery.get("steps", {})
        if not isinstance(overrides, dict) or not overrides:
            raise ValueError("Recovery overrides do not define any steps")
        session_root = path.parent.resolve()
        by_id = {step.get("step_id"): step for step in value.get("steps", [])}
        resolved: dict[str, str] = {}
        for step_id, relative_result in overrides.items():
            if step_id not in by_id:
                raise ValueError(f"Recovery override references unknown step {step_id}")
            result = (session_root / str(relative_result)).resolve()
            if not result.is_relative_to(session_root):
                raise ValueError(f"Recovery result escapes the session: {result}")
            if not (result / "summary.csv").is_file() or not (
                result / "metadata.json"
            ).is_file():
                raise ValueError(f"Recovery result is incomplete: {result}")
            step = by_id[step_id]
            step["accepted_result"] = str(result)
            step["status"] = "completed"
            step["validation"] = {
                "valid": True,
                "errors": [],
                "warnings": ["Accepted from targeted recovery override"],
            }
            resolved[step_id] = str(result)
        incomplete = [
            step for step in value.get("steps", []) if step.get("status") != "completed"
        ]
        value["completed_steps"] = len(value.get("steps", [])) - len(incomplete)
        value["failed_steps"] = len(incomplete)
        if not incomplete:
            value["state"] = "completed"
            value["message"] = "Campaign completed after targeted recovery"
        value["recovery_overrides"] = {
            "file": str(recovery_path.resolve()),
            "reason": recovery.get("reason", ""),
            "steps": resolved,
        }

    if value.get("state") != "completed":
        raise ValueError("Expected a completed campaign manifest")
    if value.get("failed_steps") or value.get("completed_steps") != len(value["steps"]):
        raise ValueError("Campaign contains failed or incomplete steps")
    return value


def _packet_rows(manifest: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, str]]]]:
    rows: list[dict[str, Any]] = []
    raw_by_direction: dict[str, list[dict[str, str]]] = {"tx": [], "rx": []}
    for step in manifest["steps"]:
        if step.get("result_kind") != "packet":
            continue
        result = Path(step["accepted_result"])
        aggregates = _read_csv(result / "aggregates.csv")
        summary = _read_csv(result / "summary.csv")
        if len(aggregates) != 1 or len(summary) != int(manifest["config"]["repetitions"]):
            raise ValueError(f"Unexpected packet matrix in {result}")
        aggregate = aggregates[0]
        direction = aggregate["measurement_direction"]
        params = json.loads(aggregate["parameters_json"])
        sample_loss = [_f(item["sample_loss_percent"]) for item in summary]
        energy = _f(aggregate["energy_total_uJ_mean"])
        energy_sd = _f(aggregate["energy_total_uJ_stdev"])
        row = {
            "profile_id": aggregate["profile_id"],
            "module": summary[0]["module"],
            "measurement_direction": direction,
            "payload_bytes": int(aggregate["payload_bytes"]),
            "tx_power_dbm": _f(params["tx_power_dbm"]),
            "spreading_factor": int(params["spreading_factor"]),
            "bandwidth_khz": _bandwidth_khz(params),
            "runs": int(aggregate["runs"]),
            "events_detected": int(aggregate["events_detected"]),
            "packets_attempted": int(aggregate["packets_attempted"]),
            "packets_received": int(aggregate["packets_received"]),
            "packets_lost": int(aggregate["packets_lost"]),
            "packet_loss_percent": _f(aggregate["packet_loss_percent"]),
            "event_duration_ms_mean": _f(aggregate["event_duration_ms_mean"]),
            "event_duration_ms_stdev": _f(aggregate["event_duration_ms_stdev"]),
            "event_mean_uA_mean": _f(aggregate["event_mean_uA_mean"]),
            "event_peak_uA_mean": _f(aggregate["event_peak_uA_mean"]),
            "energy_total_uJ_mean": energy,
            "energy_total_uJ_stdev": energy_sd,
            "energy_total_mJ_mean": energy / 1000.0,
            "energy_cv_percent": 100.0 * energy_sd / energy if energy else 0.0,
            "sample_loss_percent_mean": statistics.fmean(sample_loss),
            "sample_loss_percent_max": max(sample_loss),
            "source_directory": result.name,
        }
        rows.append(row)
        raw_by_direction[direction].extend(summary)
    rows.sort(
        key=lambda row: (
            row["measurement_direction"],
            row["tx_power_dbm"],
            row["spreading_factor"],
            row["payload_bytes"],
        )
    )
    tx = [row for row in rows if row["measurement_direction"] == "tx"]
    rx = [row for row in rows if row["measurement_direction"] == "rx"]
    sizes = {row["payload_bytes"] for row in tx}
    powers = {row["tx_power_dbm"] for row in tx}
    sfs = {row["spreading_factor"] for row in tx}
    expected_tx = {(size, power, sf) for size in sizes for power in powers for sf in sfs}
    actual_tx = {(row["payload_bytes"], row["tx_power_dbm"], row["spreading_factor"]) for row in tx}
    expected_rx = {(size, max(powers), sf) for size in sizes for sf in sfs}
    actual_rx = {(row["payload_bytes"], row["tx_power_dbm"], row["spreading_factor"]) for row in rx}
    if actual_tx != expected_tx or actual_rx != expected_rx:
        raise ValueError("Incomplete LoRa packet matrix")
    if any(row["runs"] != 5 or row["events_detected"] != 5 for row in rows):
        raise ValueError("A packet point lacks a valid energy capture")
    if any(row["sample_loss_percent_max"] > 1.0 for row in rows):
        raise ValueError("PPK2 sample loss exceeds 1%")
    return rows, raw_by_direction


def _valid_continuous_status(row: dict[str, Any]) -> bool:
    serial_errors = next(
        (
            int(part.partition("=")[2])
            for part in row.get("transmitter_response", "").split(" | ")
            if part.startswith("SERIAL_ERRORS=")
        ),
        0,
    )
    if serial_errors:
        return False
    if row["status"] == "ok":
        return True
    return (
        row["status"] == "no_rx_data"
        and row["measurement_direction"] == "rx"
        and int(row["frames_transmitted"]) > 0
        and int(row["frames_received"]) == 0
        and float(row["frame_loss_percent"]) == 100.0
        and "SERIAL_ERRORS=0" in row.get("transmitter_response", "")
    )


def _continuous_rows(manifest: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for step in manifest["steps"]:
        if step.get("result_kind") != "continuous":
            continue
        result = Path(step["accepted_result"])
        for raw in _read_csv(result / "summary.csv"):
            row: dict[str, Any] = dict(raw)
            parameters = json.loads(raw.get("parameters_json") or "{}")
            for field in (
                "voltage_mv",
                "tx_power_dbm",
                "spreading_factor",
                "bandwidth_khz",
                "frame_bytes",
                "inter_frame_gap_ms",
                "requested_duration_s",
                "actual_tx_duration_ms",
                "frames_transmitted",
                "bytes_transmitted",
                "captured_samples",
                "active_samples",
                "active_window_s",
                "sample_loss_percent",
                "baseline_mean_uA",
                "baseline_median_uA",
                "mean_current_uA",
                "stdev_current_uA",
                "peak_current_uA",
                "mean_excess_current_uA",
                "mean_power_mW",
                "stdev_power_mW",
                "peak_power_mW",
                "mean_excess_power_mW",
                "energy_60s_mJ",
            ):
                row[field] = _f(raw.get(field))
            row["bandwidth_khz"] = _bandwidth_khz(
                {**parameters, "bandwidth_khz": raw.get("bandwidth_khz")}
            )
            row["frames_received"] = int(raw["frames_received"]) if raw["frames_received"] else ""
            row["frame_loss_percent"] = _f(raw["frame_loss_percent"]) if raw["frame_loss_percent"] else ""
            row["source_directory"] = result.name
            if not _valid_continuous_status(row) or row["requested_duration_s"] != 60.0:
                raise ValueError(f"Invalid continuous row in {result}")
            if row["sample_loss_percent"] > 1.0:
                raise ValueError(f"PPK2 sample loss exceeds 1% in {result}")
            rows.append(row)
    rows.sort(key=lambda row: (row["measurement_direction"], row["spreading_factor"], row["tx_power_dbm"]))
    if len(rows) != 12:
        raise ValueError(f"Expected 12 continuous points, found {len(rows)}")
    return rows


def _loss_rows(continuous: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source in continuous:
        if source["measurement_direction"] != "rx":
            continue
        transmitted = int(source["frames_transmitted"])
        received = int(source["frames_received"])
        duration = float(source["requested_duration_s"])
        frame_bytes = int(source["frame_bytes"])
        loss = float(source["frame_loss_percent"])
        rows.append(
            {
                "spreading_factor": int(source["spreading_factor"]),
                "bandwidth_khz": source["bandwidth_khz"],
                "effective_payload_rate_kbps": transmitted * frame_bytes * 8.0 / duration / 1000.0,
                "tx_power_dbm": source["tx_power_dbm"],
                "frames_transmitted": transmitted,
                "frames_received": received,
                "frames_lost": transmitted - received,
                "frame_loss_percent": loss,
                "delivery_percent": 100.0 - loss,
                "requested_duration_s": duration,
                "frame_bytes": frame_bytes,
                "inter_frame_gap_ms": int(source["inter_frame_gap_ms"]),
                "status": source["status"],
                "source_directory": source["source_directory"],
            }
        )
    rows.sort(key=lambda row: (row["tx_power_dbm"], -row["effective_payload_rate_kbps"]))
    powers = sorted({row["tx_power_dbm"] for row in rows})
    spreading_factors = sorted({row["spreading_factor"] for row in rows})
    if len(powers) != 3 or len(spreading_factors) != 3:
        raise ValueError("Expected three TX powers and three spreading factors")
    expected = {(power, sf) for power in powers for sf in spreading_factors}
    actual = {(row["tx_power_dbm"], row["spreading_factor"]) for row in rows}
    if actual != expected:
        raise ValueError("Incomplete LoRa loss matrix")
    return rows


def _write_csv(path: Path, fields: tuple[str, ...] | list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _style(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(55, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _sheet(workbook: Workbook, name: str, fields: list[str] | tuple[str, ...], rows: list[dict[str, Any]]):
    sheet = workbook.create_sheet(name)
    sheet.append(list(fields))
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    _style(sheet)
    return sheet


def _write_packet_xlsx(path: Path, aggregate: list[dict[str, Any]], runs: list[dict[str, str]], manifest: dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    _sheet(workbook, "aggregates", PACKET_FIELDS, aggregate)
    run_fields = list(runs[0])
    _sheet(workbook, "raw_runs", run_fields, runs)
    metadata = workbook.create_sheet("campaign")
    metadata.append(["field", "value"])
    for field in ("state", "started_utc", "finished_utc", "completed_steps", "failed_steps"):
        metadata.append([field, manifest.get(field, "")])
    for field, value in manifest["config"].items():
        metadata.append([f"config.{field}", value])
    _style(metadata)
    workbook.save(path)


def _write_continuous_xlsx(path: Path, rows: list[dict[str, Any]], loss: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    fields = list(rows[0])
    _sheet(workbook, "continuous_results", fields, rows)
    _sheet(workbook, "loss_results", LOSS_FIELDS, loss)
    matrix = workbook.create_sheet("loss_matrix_percent")
    spreading_factors = sorted({int(row["spreading_factor"]) for row in loss})
    matrix.append(["TX power [dBm]", *(f"SF{sf}" for sf in spreading_factors)])
    by_key = {(row["tx_power_dbm"], row["spreading_factor"]): row for row in loss}
    for power in sorted({row["tx_power_dbm"] for row in loss}):
        matrix.append(
            [
                power,
                *(by_key[(power, sf)]["frame_loss_percent"] for sf in spreading_factors),
            ]
        )
    _style(matrix)
    workbook.save(path)


def _energy_coordinates(rows: list[dict[str, Any]]) -> str:
    return " ".join(
        f"({_n(row['payload_bytes'])},{_n(row['energy_total_mJ_mean'])}) +- (0,{_n(row['energy_total_uJ_stdev']/1000.0)})"
        for row in sorted(rows, key=lambda item: item["payload_bytes"])
    )


def _tex_header(title: str, panels: int, ymin: float, ymax: float) -> list[str]:
    return [
        r"\documentclass[tikz,border=6pt]{standalone}", r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}", r"\usepackage{pgfplots}",
        r"\usepgfplotslibrary{groupplots}", r"\usetikzlibrary{calc}",
        r"\pgfplotsset{compat=1.18}", r"\begin{document}", r"\begin{tikzpicture}",
        rf"\begin{{groupplot}}[/pgf/number format/use comma, group style={{group size={panels} by 1, horizontal sep=1.25cm}},",
        rf"  width=5.8cm,height=6.2cm,xmode=log,log basis x=2,ymode=log,log basis y=10,ymin={_n(ymin)},ymax={_n(ymax)},",
        r"  xtick={8,32,128},xticklabels={8,32,128},grid=both,xlabel={Payload [B]},",
        r"  legend style={draw=none,fill=none,font=\small,at={(0.5,-0.27)},anchor=north,legend columns=-1}]",
        rf"% {title}",
    ]


def _energy_limits(rows: list[dict[str, Any]]) -> tuple[float, float]:
    values = [row["energy_total_mJ_mean"] for row in rows]
    return 10 ** math.floor(math.log10(min(values) * 0.75)), 10 ** math.ceil(math.log10(max(values) * 1.25))


def _tex_footer(module: str, title: str, panels: int, note: str) -> list[str]:
    center = (panels + 1) // 2
    return [
        r"\end{groupplot}",
        rf"\node[font=\bfseries\large] at ($(group c{center}r1.north)+(0,0.85cm)$) {{{module}: {title}}};",
        rf"\node[font=\footnotesize,align=center,text width=18cm] at ($(group c{center}r1.south)+(0,-2.75cm)$) {{{note}}};",
        r"\end{tikzpicture}", r"\end{document}", "",
    ]


def _write_tx_energy_tex(path: Path, tx: list[dict[str, Any]], module: str) -> None:
    ymin, ymax = _energy_limits(tx)
    powers = sorted({row["tx_power_dbm"] for row in tx})
    spreading_factors = sorted({int(row["spreading_factor"]) for row in tx})
    lines = _tex_header("TX energy", len(powers), ymin, ymax)
    for power in powers:
        ylabel = r",ylabel={Mean total energy [mJ]}" if power == powers[0] else ""
        lines.append(rf"\nextgroupplot[title={{{_latex_n(power)} dBm}}{ylabel}]")
        for index, sf in enumerate(spreading_factors):
            selected = [row for row in tx if row["tx_power_dbm"] == power and row["spreading_factor"] == sf]
            style = index % len(COLORS)
            lines.append(rf"\addplot+[{COLORS[style]},{STYLES[style]},mark={MARKERS[style]},error bars/.cd,y dir=both,y explicit] coordinates {{{_energy_coordinates(selected)}}};")
            if power == powers[len(powers) // 2]:
                lines.append(rf"\addlegendentry{{SF{sf}}}")
    lines.extend(_tex_footer(module, "TX energy", len(powers), "Five repetitions per point; error bars show standard deviation. Both axes are logarithmic."))
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_tx_rx_tex(path: Path, packet: list[dict[str, Any]], module: str) -> None:
    power = max(row["tx_power_dbm"] for row in packet)
    selected_all = [row for row in packet if row["tx_power_dbm"] == power]
    spreading_factors = sorted({int(row["spreading_factor"]) for row in selected_all})
    ymin, ymax = _energy_limits(selected_all)
    lines = _tex_header("TX--RX energy", len(spreading_factors), ymin, ymax)
    for index, sf in enumerate(spreading_factors):
        lines.append(rf"\nextgroupplot[title={{SF{sf}}}{',ylabel={Mean total energy [mJ]}' if index == 0 else ''}]")
        tx = [row for row in selected_all if row["measurement_direction"] == "tx" and row["spreading_factor"] == sf]
        rx = [row for row in selected_all if row["measurement_direction"] == "rx" and row["spreading_factor"] == sf]
        lines.append(rf"\addplot+[blue!75!black,solid,mark=*,error bars/.cd,y dir=both,y explicit] coordinates {{{_energy_coordinates(tx)}}};")
        if index == len(spreading_factors) // 2:
            lines.append(r"\addlegendentry{TX}")
        lines.append(rf"\addplot+[red!75!black,dashed,mark=square*,error bars/.cd,y dir=both,y explicit] coordinates {{{_energy_coordinates(rx)}}};")
        if index == len(spreading_factors) // 2:
            lines.append(r"\addlegendentry{RX}")
    lines.extend(_tex_footer(module, f"TX--RX energy at {_latex_n(power)} dBm", len(spreading_factors), "RX integrates one deterministic LoRa airtime window; five repetitions per point."))
    path.write_text("\n".join(lines), encoding="utf-8")


def _coordinates(rows: list[dict[str, Any]], x: str, y: str) -> str:
    return " ".join(f"({_n(row[x])},{_n(row[y])})" for row in rows)


def _scaled_coordinates(rows: list[dict[str, Any]], x: str, y: str, divisor: float) -> str:
    return " ".join(f"({_n(row[x])},{_n(row[y] / divisor)})" for row in rows)


def _write_continuous_power_tex(path: Path, rows: list[dict[str, Any]], module: str) -> None:
    tx = sorted((row for row in rows if row["measurement_direction"] == "tx"), key=lambda row: row["tx_power_dbm"])
    sf = int(tx[0]["spreading_factor"])
    rx = sorted((row for row in rows if row["measurement_direction"] == "rx" and int(row["spreading_factor"]) == sf), key=lambda row: row["tx_power_dbm"])
    powers = sorted(row["tx_power_dbm"] for row in tx)
    ticks = ",".join(_n(power) for power in powers)
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}", r"\usepackage{pgfplots}",
        r"\usepgfplotslibrary{groupplots}", r"\usetikzlibrary{calc}", r"\pgfplotsset{compat=1.18}", r"\begin{document}", r"\begin{tikzpicture}",
        rf"\begin{{groupplot}}[group style={{group size=2 by 1,horizontal sep=1.5cm}},width=8cm,height=6.2cm,xtick={{{ticks}}},xlabel={{Configured TX power [dBm]}},grid=both,legend style={{draw=none,at={{(0.5,-0.23)}},anchor=north,legend columns=2}}]",
        r"\nextgroupplot[title={Total power},ylabel={Mean electrical power [mW]}]",
        rf"\addplot+[blue!75!black,mark=*] coordinates {{{_coordinates(tx, 'tx_power_dbm', 'mean_power_mW')}}};\addlegendentry{{TX}}",
        rf"\addplot+[red!75!black,dashed,mark=square*] coordinates {{{_coordinates(rx, 'tx_power_dbm', 'mean_power_mW')}}};\addlegendentry{{RX}}",
        r"\nextgroupplot[title={Mean current},ylabel={Mean current [mA]}]",
        rf"\addplot+[blue!75!black,mark=*] coordinates {{{_scaled_coordinates(tx, 'tx_power_dbm', 'mean_current_uA', 1000.0)}}};\addlegendentry{{TX}}",
        rf"\addplot+[red!75!black,dashed,mark=square*] coordinates {{{_scaled_coordinates(rx, 'tx_power_dbm', 'mean_current_uA', 1000.0)}}};\addlegendentry{{RX}}",
        r"\end{groupplot}", rf"\node[font=\bfseries\large] at ($(group c1r1.north)!0.5!(group c2r1.north)+(0,0.9cm)$) {{{module}: continuous average power (SF{sf})}};",
        r"\end{tikzpicture}", r"\end{document}", "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_loss_tex(path: Path, rows: list[dict[str, Any]], module: str) -> None:
    rates = [row["effective_payload_rate_kbps"] for row in rows]
    spreading_factors = sorted({int(row["spreading_factor"]) for row in rows})
    rates_by_sf = {
        sf: statistics.fmean(
            row["effective_payload_rate_kbps"]
            for row in rows
            if row["spreading_factor"] == sf
        )
        for sf in spreading_factors
    }
    tick_rates = [rates_by_sf[sf] for sf in reversed(spreading_factors)]
    ticks = ",".join(_n(rate) for rate in tick_rates)
    tick_labels = ",".join(f"{rate:.3g}" for rate in tick_rates)
    xmin, xmax = min(rates) * 0.75, max(rates) * 1.25
    ymax = min(100, max(10, math.ceil(max(row["frame_loss_percent"] for row in rows) * 1.25 / 5) * 5))
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}", r"\usepackage{pgfplots}", r"\pgfplotsset{compat=1.18}",
        r"\begin{document}", r"\begin{tikzpicture}", r"\begin{axis}[width=12.5cm,height=7.4cm,",
        rf" title={{{module}: loss versus effective payload speed}},xmode=log,xmin={_n(xmin)},xmax={_n(xmax)},ymin=0,ymax={_n(ymax)},",
        rf" xtick={{{ticks}}},xticklabels={{{tick_labels}}},",
        r" xlabel={Effective payload speed [kbps]},ylabel={Lost frames [\%]},grid=both,legend style={draw=none,at={(0.03,0.97)},anchor=north west}]",
    ]
    powers = sorted({row["tx_power_dbm"] for row in rows}, reverse=True)
    for index, power in enumerate(powers):
        selected = sorted((row for row in rows if row["tx_power_dbm"] == power), key=lambda row: row["effective_payload_rate_kbps"])
        style = index % len(COLORS)
        lines.append(rf"\addplot+[{COLORS[style]},{STYLES[style]},mark={MARKERS[style]}] coordinates {{{_coordinates(selected, 'effective_payload_rate_kbps', 'frame_loss_percent')}}};")
        lines.append(rf"\addlegendentry{{{_latex_n(power)} dBm}}")
    mapping = ", ".join(
        f"SF{sf}: {rates_by_sf[sf]:.3g} kbps"
        for sf in spreading_factors
    )
    lines.extend([r"\end{axis}", rf"\node[font=\footnotesize,align=center,text width=12cm] at (6.25,-1.45) {{60 s per point, 64 B frames, 15 ms gap; {mapping}.}};", r"\end{tikzpicture}", r"\end{document}", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def _copy_provenance(manifest_path: Path, output: Path) -> None:
    destination = output / "campaign_logs"
    destination.mkdir(parents=True, exist_ok=True)
    session = manifest_path.parent
    for name in (
        "manifest.json",
        "session.log",
        "codex_callback.log",
        "recovery_overrides.json",
    ):
        source = session / name
        if source.exists():
            shutil.copy2(source, destination / name)
    if (session / "logs").exists():
        shutil.copytree(session / "logs", destination / "attempts", dirs_exist_ok=True)
    recovery = session / "recovery"
    if recovery.exists():
        for source in recovery.rglob("*"):
            if source.is_file() and (
                source.suffix == ".log"
                or source.name
                in {
                    "metadata.json",
                    "summary.csv",
                    "aggregates.csv",
                    "manifest.json",
                    "session.log",
                    "codex_callback.log",
                }
            ):
                target = destination / "recovery" / source.relative_to(recovery)
                target.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(source, target)


def generate(
    manifest_path: Path,
    output: Path,
    base_name: str,
    module_name: str | None = None,
) -> list[Path]:
    manifest = _manifest(manifest_path)
    packet, raw = _packet_rows(manifest)
    continuous = _continuous_rows(manifest)
    if module_name:
        for row in packet:
            row["module"] = module_name
        for row in continuous:
            row["module"] = module_name
    loss = _loss_rows(continuous)
    output.mkdir(parents=True, exist_ok=True)
    base = output / base_name
    tx = [row for row in packet if row["measurement_direction"] == "tx"]
    rx = [row for row in packet if row["measurement_direction"] == "rx"]
    _write_csv(base.with_name(base.name + "_data").with_suffix(".csv"), PACKET_FIELDS, packet)
    _write_csv(base.with_name(base.name + "_tx").with_suffix(".csv"), PACKET_FIELDS, tx)
    _write_csv(base.with_name(base.name + "_rx").with_suffix(".csv"), PACKET_FIELDS, rx)
    _write_packet_xlsx(base.with_name(base.name + "_tx").with_suffix(".xlsx"), tx, raw["tx"], manifest)
    _write_packet_xlsx(base.with_name(base.name + "_rx").with_suffix(".xlsx"), rx, raw["rx"], manifest)
    continuous_fields = list(continuous[0])
    _write_csv(base.with_name(base.name + "_continuous").with_suffix(".csv"), continuous_fields, continuous)
    _write_continuous_xlsx(base.with_name(base.name + "_continuous").with_suffix(".xlsx"), continuous, loss)
    _write_csv(base.with_name(base.name + "_loss_vs_speed").with_suffix(".csv"), LOSS_FIELDS, loss)
    loss_book = Workbook()
    loss_book.remove(loss_book.active)
    _sheet(loss_book, "loss_results", LOSS_FIELDS, loss)
    loss_book.save(base.with_name(base.name + "_loss_vs_speed").with_suffix(".xlsx"))
    module = packet[0]["module"]
    _write_tx_energy_tex(base.with_name(base.name + "_tx_energy").with_suffix(".tex"), tx, module)
    _write_tx_rx_tex(base.with_name(base.name + "_tx_rx_energy").with_suffix(".tex"), packet, module)
    _write_continuous_power_tex(base.with_name(base.name + "_continuous_average_power").with_suffix(".tex"), continuous, module)
    _write_loss_tex(base.with_name(base.name + "_loss_vs_speed").with_suffix(".tex"), loss, module)
    audit = {
        "authoritative_session": str(manifest_path.parent.resolve()),
        "packet_points": len(packet),
        "packet_runs": len(raw["tx"]) + len(raw["rx"]),
        "continuous_points": len(continuous),
        "maximum_energy_cv_percent": max(row["energy_cv_percent"] for row in packet),
        "maximum_ppk_sample_loss_percent": max(row["sample_loss_percent_max"] for row in packet),
        "packet_losses": sum(row["packets_lost"] for row in packet),
        "recovery_overrides": manifest.get("recovery_overrides", {}),
        "diagnostic_recovery_summaries": [
            str(path.relative_to(manifest_path.parent))
            for path in sorted((manifest_path.parent / "recovery").rglob("summary.csv"))
        ]
        if (manifest_path.parent / "recovery").exists()
        else [],
        "notes": [
            "RX energy uses a deterministic integration window equal to calculated LoRa airtime.",
            "Verified total radio loss is retained when the transmitter completed with zero serial errors.",
            "Loss results are preserved as measured; isolated packet losses were not cherry-picked away.",
        ],
    }
    base.with_name(base.name + "_audit.json").write_text(json.dumps(audit, indent=2) + "\n", encoding="utf-8")
    _copy_provenance(manifest_path, output)
    from render_lora_campaign_plots import render as render_plots

    render_plots(output, base_name)
    return sorted(path for path in output.rglob("*") if path.is_file())


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate consolidated LoRa campaign reports")
    parser.add_argument("manifest", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--base-name", required=True)
    parser.add_argument("--module-name")
    args = parser.parse_args()
    for output in generate(
        args.manifest,
        args.output_dir,
        args.base_name,
        module_name=args.module_name,
    ):
        print(output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
