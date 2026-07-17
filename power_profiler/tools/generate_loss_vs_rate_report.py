from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


RATES = (1.2, 38.4, 250.0)
POWERS = (-30.0, 0.0, 10.0)
OUTPUT_FIELDS = (
    "bit_rate_kbps",
    "tx_power_dbm",
    "frames_transmitted",
    "frames_received",
    "frames_lost",
    "frame_loss_percent",
    "delivery_percent",
    "requested_duration_s",
    "frame_bytes",
    "inter_frame_gap_ms",
    "status",
    "source_directory",
)


def read_results(result_dirs: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result_dir in result_dirs:
        metadata = json.loads((result_dir / "metadata.json").read_text(encoding="utf-8"))
        if metadata.get("measurement_direction") != "rx":
            raise ValueError(f"Not an RX result directory: {result_dir}")
        with (result_dir / "summary.csv").open(encoding="utf-8", newline="") as stream:
            for raw in csv.DictReader(stream):
                transmitted = int(float(raw["frames_transmitted"]))
                received = int(raw["frames_received"])
                loss = float(raw["frame_loss_percent"])
                rows.append(
                    {
                        "bit_rate_kbps": float(raw["bit_rate_kbps"]),
                        "tx_power_dbm": float(raw["tx_power_dbm"]),
                        "frames_transmitted": transmitted,
                        "frames_received": received,
                        "frames_lost": transmitted - received,
                        "frame_loss_percent": loss,
                        "delivery_percent": 100.0 - loss,
                        "requested_duration_s": float(raw["requested_duration_s"]),
                        "frame_bytes": int(raw["frame_bytes"]),
                        "inter_frame_gap_ms": int(raw["inter_frame_gap_ms"]),
                        "status": raw["status"],
                        "source_directory": str(result_dir.resolve()),
                    }
                )
    rows.sort(key=lambda row: (row["tx_power_dbm"], row["bit_rate_kbps"]))
    validate(rows)
    return rows


def validate(rows: list[dict[str, Any]]) -> None:
    expected = {(power, rate) for power in POWERS for rate in RATES}
    actual = {(row["tx_power_dbm"], row["bit_rate_kbps"]) for row in rows}
    if actual != expected or len(rows) != len(expected):
        raise ValueError(f"Incomplete or duplicate loss matrix: {sorted(actual)}")
    for row in rows:
        if row["requested_duration_s"] != 60.0:
            raise ValueError("All measurements must use 60-second windows")
        if row["frame_bytes"] != 32 or row["inter_frame_gap_ms"] != 15:
            raise ValueError("All measurements must use 32-byte frames and a 15 ms gap")
        if not 0.0 <= row["frame_loss_percent"] <= 100.0:
            raise ValueError(f"Invalid packet loss: {row}")


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def _style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")


def _fit_columns(sheet) -> None:
    for index, column in enumerate(sheet.columns, start=1):
        width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    results = workbook.active
    results.title = "loss_results"
    results.append(OUTPUT_FIELDS)
    for row in rows:
        results.append([row[field] for field in OUTPUT_FIELDS])
    _style_header(results)
    results.freeze_panes = "A2"
    results.auto_filter.ref = results.dimensions
    _fit_columns(results)

    matrix = workbook.create_sheet("loss_matrix_percent")
    matrix.append(["TX power [dBm]", *RATES])
    by_key = {
        (row["tx_power_dbm"], row["bit_rate_kbps"]): row
        for row in rows
    }
    for power in POWERS:
        matrix.append(
            [power, *(by_key[(power, rate)]["frame_loss_percent"] for rate in RATES)]
        )
    _style_header(matrix)
    matrix.freeze_panes = "B2"
    _fit_columns(matrix)
    workbook.save(path)


def _coordinates(rows: list[dict[str, Any]], power: float) -> str:
    selected = [row for row in rows if row["tx_power_dbm"] == power]
    return " ".join(
        f"({row['bit_rate_kbps']:g},{row['frame_loss_percent']:.6f})"
        for row in selected
    )


def write_tex(path: Path, rows: list[dict[str, Any]], module_title: str) -> None:
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{axis}[width=12.5cm, height=7.4cm,",
        rf"  title={{{module_title}: packet loss în funcție de viteza radio}},",
        r"  xmode=log, log basis x=10,",
        r"  xmin=0.8, xmax=400, xtick={1.2,38.4,250},",
        r"  xticklabels={1{,}2,38{,}4,250},",
        r"  ymin=0, ymax=105, ytick={0,20,40,60,80,100},",
        r"  xlabel={Viteza radio [kbps]}, ylabel={Cadre pierdute [\%]},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"  every axis plot/.append style={line width=1.25pt, mark size=2.8pt},",
        r"  legend style={draw=none, fill=white, fill opacity=0.85, text opacity=1,",
        r"    at={(0.03,0.97)}, anchor=north west, legend columns=1},",
        r"]",
        rf"\addplot+[green!50!black, mark=triangle*] coordinates {{{_coordinates(rows, 10.0)}}};",
        r"\addlegendentry{+10 dBm}",
        rf"\addplot+[orange!90!black, dashed, mark=square*] coordinates {{{_coordinates(rows, 0.0)}}};",
        r"\addlegendentry{0 dBm}",
        rf"\addplot+[red!75!black, dashdotted, mark=*] coordinates {{{_coordinates(rows, -30.0)}}};",
        r"\addlegendentry{$-30$ dBm}",
        r"\end{axis}",
        r"\node[font=\footnotesize, align=center, text width=12cm, anchor=north]",
        r"  at ([yshift=-1.45cm]current axis.south)",
        r"  {60 s/punct; cadre de 32 B; pauză de 15 ms; puterea din legendă este puterea configurată a emițătorului.};",
        r"\end{tikzpicture}",
        r"\end{document}",
        "",
    ]
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate CC1101 packet-loss versus rate report")
    parser.add_argument("result_dirs", nargs=3, type=Path)
    parser.add_argument("--output-base", required=True, type=Path)
    parser.add_argument("--module-title", default="CC1101 V2, 868 MHz")
    args = parser.parse_args()

    rows = read_results(args.result_dirs)
    args.output_base.parent.mkdir(parents=True, exist_ok=True)
    csv_path = args.output_base.with_suffix(".csv")
    xlsx_path = args.output_base.with_suffix(".xlsx")
    tex_path = args.output_base.with_suffix(".tex")
    write_csv(csv_path, rows)
    write_xlsx(xlsx_path, rows)
    write_tex(tex_path, rows, args.module_title)
    for output in (csv_path, xlsx_path, tex_path):
        print(output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
