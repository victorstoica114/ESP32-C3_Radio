from __future__ import annotations

import argparse
import copy
import csv
import json
import math
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import generate_continuous_report as continuous_report
from generate_transfer_report import (
    REPORT_FIELDS,
    build_report,
    write_csv as write_transfer_csv,
    write_xlsx as write_transfer_xlsx,
)


COLORS = (
    "blue!75!black",
    "orange!90!black",
    "green!50!black",
    "red!75!black",
    "violet!75!black",
)
LINE_STYLES = ("solid", "dashed", "dashdotted", "densely dotted")
MARKERS = ("*", "square*", "triangle*", "diamond*")
LOSS_FIELDS = (
    "bit_rate_kbps",
    "rf_profile",
    "tx_power_dbm",
    "frames_transmitted",
    "frames_received",
    "frames_lost",
    "frame_loss_percent",
    "delivery_percent",
    "requested_duration_s",
    "frame_bytes",
    "inter_frame_gap_ms",
    "status",
    "source_directory",
)


def _number(value: float | int) -> str:
    return f"{value:.9g}"


def _latex_number(value: float | int) -> str:
    return _number(value).replace(".", "{,}")


def _setting_key(row: dict[str, Any]) -> str:
    profile = str(row.get("rf_profile") or "")
    return profile or f"rate:{_number(float(row['bit_rate_kbps']))}"


def _radio_settings(rows: list[dict[str, Any]]) -> list[tuple[str, float, str]]:
    settings: dict[str, tuple[float, str]] = {}
    for row in rows:
        key = _setting_key(row)
        rate = float(row["bit_rate_kbps"])
        profile = str(row.get("rf_profile") or "")
        label = f"{profile} ({_number(rate)} kbps)" if profile else f"{_number(rate)} kbps"
        settings[key] = (rate, label)
    return [
        (key, rate, label)
        for key, (rate, label) in sorted(
            settings.items(), key=lambda item: (item[1][0], item[0])
        )
    ]


def _read_manifest(path: Path) -> dict[str, Any]:
    manifest = copy.deepcopy(json.loads(path.read_text(encoding="utf-8")))
    if manifest.get("kind") != "campaign":
        raise ValueError("Expected a campaign manifest")

    recovery_path = path.parent / "recovery_overrides.json"
    if recovery_path.exists():
        recovery = json.loads(recovery_path.read_text(encoding="utf-8"))
        overrides = recovery.get("steps", {})
        if not isinstance(overrides, dict) or not overrides:
            raise ValueError("Recovery overrides do not define any steps")
        session_root = path.parent.resolve()
        by_id = {step.get("step_id"): step for step in manifest.get("steps", [])}
        resolved: dict[str, str] = {}
        for step_id, relative_result in overrides.items():
            if step_id not in by_id:
                raise ValueError(f"Recovery override references unknown step {step_id}")
            result = (session_root / str(relative_result)).resolve()
            if not result.is_relative_to(session_root):
                raise ValueError(f"Recovery result escapes the session: {result}")
            if not (result / "summary.csv").is_file() or not (
                result / "metadata.json"
            ).is_file():
                raise ValueError(f"Recovery result is incomplete: {result}")
            step = by_id[step_id]
            step["accepted_result"] = str(result)
            step["status"] = "completed"
            step["validation"] = {
                "valid": True,
                "errors": [],
                "warnings": ["Accepted from targeted recovery override"],
            }
            resolved[step_id] = str(result)
        incomplete = [
            step
            for step in manifest.get("steps", [])
            if step.get("status") != "completed"
        ]
        manifest["completed_steps"] = len(manifest.get("steps", [])) - len(incomplete)
        manifest["failed_steps"] = len(incomplete)
        if not incomplete:
            manifest["state"] = "completed"
            manifest["message"] = "Campaign completed after targeted recovery"
        manifest["recovery_overrides"] = {
            "file": str(recovery_path.resolve()),
            "reason": recovery.get("reason", ""),
            "steps": resolved,
        }

    if manifest.get("state") != "completed":
        raise ValueError("The manifest is not a completed campaign")
    if manifest.get("failed_steps") != 0:
        raise ValueError("The campaign contains failed steps")
    if manifest.get("completed_steps") != len(manifest.get("steps", [])):
        raise ValueError("The campaign step count is incomplete")
    return manifest


def _collect_packet_results(
    manifest: dict[str, Any], direction: str
) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, Any]]:
    report: list[dict[str, Any]] = []
    summary: list[dict[str, str]] = []
    metadata: dict[str, Any] | None = None
    prefix = direction + "_"
    for step in manifest["steps"]:
        if step.get("result_kind") != "packet" or not step["step_id"].startswith(prefix):
            continue
        result_dir = Path(step["accepted_result"])
        rows, run_rows, item_metadata = build_report(result_dir)
        if len(rows) != 1:
            raise ValueError(f"Expected one aggregate in {result_dir}, found {len(rows)}")
        if rows[0]["measurement_direction"] != direction:
            raise ValueError(f"Unexpected direction in {result_dir}")
        report.extend(rows)
        summary.extend(run_rows)
        metadata = metadata or item_metadata
    if metadata is None:
        raise ValueError(f"No accepted {direction.upper()} packet results")
    report.sort(
        key=lambda row: (
            row["tx_power_dbm"],
            row["bit_rate_kbps"],
            row.get("rf_profile", ""),
            row["payload_bytes"],
        )
    )
    return report, summary, metadata


def _validate_energy_matrix(
    tx_rows: list[dict[str, Any]],
    rx_rows: list[dict[str, Any]],
    repetitions: int,
) -> tuple[list[int], list[float], list[tuple[str, float, str]], float, str, int]:
    module = str(tx_rows[0]["module"])
    if any(row["module"] != module for row in tx_rows + rx_rows):
        raise ValueError("TX and RX data refer to different modules")
    sizes = sorted({int(row["payload_bytes"]) for row in tx_rows})
    powers = sorted({float(row["tx_power_dbm"]) for row in tx_rows})
    settings = _radio_settings(tx_rows)
    setting_keys = {item[0] for item in settings}
    rx_setting_keys = {_setting_key(row) for row in rx_rows}
    rx_powers = sorted({float(row["tx_power_dbm"]) for row in rx_rows})
    if len(rx_powers) != 1 or rx_powers[0] != max(powers):
        raise ValueError("RX must use the maximum tested transmitter power")
    if rx_setting_keys != setting_keys:
        raise ValueError("TX and RX cover different radio settings")
    tx_expected = {
        (size, power, setting)
        for size in sizes
        for power in powers
        for setting in setting_keys
    }
    rx_expected = {
        (size, rx_powers[0], setting) for size in sizes for setting in setting_keys
    }
    tx_actual = {
        (row["payload_bytes"], row["tx_power_dbm"], _setting_key(row))
        for row in tx_rows
    }
    rx_actual = {
        (row["payload_bytes"], row["tx_power_dbm"], _setting_key(row))
        for row in rx_rows
    }
    if tx_actual != tx_expected or len(tx_rows) != len(tx_expected):
        raise ValueError("Incomplete or duplicate TX energy matrix")
    if rx_actual != rx_expected or len(rx_rows) != len(rx_expected):
        raise ValueError("Incomplete or duplicate RX energy matrix")
    if any(
        row["runs"] != repetitions or row["status_ok_runs"] != repetitions
        for row in tx_rows + rx_rows
    ):
        raise ValueError("Every energy point must contain all valid repetitions")
    frame_limit = max(int(row["max_frame_payload_bytes"]) for row in tx_rows)
    return sizes, powers, settings, rx_powers[0], module, frame_limit


def _campaign_metadata(
    metadata: dict[str, Any],
    direction: str,
    sizes: list[int],
    powers: list[float],
    settings: list[tuple[str, float, str]],
) -> dict[str, Any]:
    merged = copy.deepcopy(metadata)
    merged["measurement_direction"] = direction
    merged["profile"]["payload_sizes"] = sizes
    for axis in merged["profile"]["axes"]:
        if axis["name"] == "tx_power_dbm":
            axis["values"] = powers
        elif axis["name"] == "bit_rate_kbps":
            axis["values"] = sorted({setting[1] for setting in settings})
        elif axis["name"] == "rf_profile":
            axis["values"] = [setting[0] for setting in settings if not setting[0].startswith("rate:")]
    return merged


def _write_energy_data(path: Path, rows: list[dict[str, Any]]) -> None:
    fields = (
        "module",
        "measurement_direction",
        "payload_bytes",
        "tx_power_dbm",
        "bit_rate_kbps",
        "rf_profile",
        "runs",
        "status_ok_runs",
        "packets_received",
        "energy_total_mJ_mean",
        "energy_total_uJ_stdev",
        "event_mean_uA_mean",
        "event_peak_uA_mean",
        "event_duration_ms_mean",
        "sample_loss_percent_max",
    )
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def _y_limits(rows: list[dict[str, Any]]) -> tuple[float, float]:
    values = [float(row["energy_total_mJ_mean"]) for row in rows]
    return (
        10 ** math.floor(math.log10(min(values) * 0.75)),
        10 ** math.ceil(math.log10(max(values) * 1.25)),
    )


def _energy_coordinates(rows: list[dict[str, Any]]) -> list[str]:
    lines = [r"  error bars/.cd, y dir=both, y explicit,", r"] coordinates {"]
    for row in sorted(rows, key=lambda item: item["payload_bytes"]):
        lines.append(
            f"  ({_number(row['payload_bytes'])},{_number(row['energy_total_mJ_mean'])}) "
            f"+- (0,{_number(float(row['energy_total_uJ_stdev']) / 1000.0)})"
        )
    lines.append(r"};")
    return lines


def _energy_document_start(
    rows: list[dict[str, Any]], sizes: list[int], groups: int
) -> list[str]:
    ymin, ymax = _y_limits(rows)
    ticks = ",".join(_number(size) for size in sizes)
    width = 5.7 if groups >= 3 else 7.2
    return [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\usepgfplotslibrary{groupplots}",
        r"\usetikzlibrary{calc}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{groupplot}[/pgf/number format/use comma,",
        rf"  group style={{group size={groups} by 1, horizontal sep=1.25cm}},",
        rf"  width={width}cm, height=6.2cm,",
        r"  xmode=log, log basis x={2}, ymode=log, log basis y={10},",
        rf"  ymin={_number(ymin)}, ymax={_number(ymax)},",
        rf"  xtick={{{ticks}}}, xticklabels={{{ticks}}},",
        r"  x tick label style={rotate=45, anchor=east, font=\small},",
        r"  y tick label style={font=\small},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"  xlabel={Logical transfer size [B]},",
        r"  every axis plot/.append style={line width=1.05pt},",
        r"  error bars/error bar style={line width=0.35pt},",
        r"  legend style={font=\small, draw=none, fill=none, at={(0.5,-0.29)},",
        r"    anchor=north, legend columns=-1},",
        r"]",
    ]


def _energy_document_end(module: str, title: str, note: str, groups: int) -> list[str]:
    center = (groups + 1) // 2
    return [
        r"\end{groupplot}",
        rf"\node[font=\bfseries\large] at ($(group c{center}r1.north)+(0,0.85cm)$)",
        rf"  {{{module}: {title}}};",
        r"\node[font=\footnotesize, align=center, text width=18.5cm]",
        rf"  at ($(group c{center}r1.south)+(0,-3.15cm)$)",
        rf"  {{{note}}};",
        r"\end{tikzpicture}",
        r"\end{document}",
        "",
    ]


def _write_tx_tex(
    path: Path,
    tx_rows: list[dict[str, Any]],
    sizes: list[int],
    powers: list[float],
    settings: list[tuple[str, float, str]],
    module: str,
    frame_limit: int,
    title: str,
) -> None:
    lines = _energy_document_start(tx_rows, sizes, len(powers))
    legend_power = powers[len(powers) // 2]
    for power in powers:
        ylabel = r", ylabel={Mean total energy [mJ]}" if power == powers[0] else ""
        lines.append(
            rf"\nextgroupplot[title={{\(P_{{\mathrm{{TX}}}}={_latex_number(power)}\,\mathrm{{dBm}}\)}}{ylabel}]"
        )
        for index, (setting, _rate, label) in enumerate(settings):
            points = [
                row for row in tx_rows
                if row["tx_power_dbm"] == power and _setting_key(row) == setting
            ]
            lines.extend(
                [
                    r"\addplot+[",
                    rf"  color={COLORS[index % len(COLORS)]}, {LINE_STYLES[index % len(LINE_STYLES)]}, mark={MARKERS[index % len(MARKERS)]}, mark size=2.2pt,",
                    *_energy_coordinates(points),
                ]
            )
            if power == legend_power:
                lines.append(rf"\addlegendentry{{{label}}}")
    note = (
        f"Mean of 5 repetitions; error bars show standard deviation. Logical transfers "
        f"above {frame_limit} B are fragmented into physical frames of at most "
        f"{frame_limit} B; both axes are logarithmic."
    )
    lines.extend(_energy_document_end(module, title, note, len(powers)))
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_tx_rx_tex(
    path: Path,
    tx_rows: list[dict[str, Any]],
    rx_rows: list[dict[str, Any]],
    sizes: list[int],
    settings: list[tuple[str, float, str]],
    power: float,
    module: str,
) -> None:
    plotted = [row for row in tx_rows if row["tx_power_dbm"] == power] + rx_rows
    lines = _energy_document_start(plotted, sizes, len(settings))
    legend_setting = settings[len(settings) // 2][0]
    for index, (setting, _rate, label) in enumerate(settings):
        ylabel = r", ylabel={Mean total energy [mJ]}" if index == 0 else ""
        lines.append(rf"\nextgroupplot[title={{{label}}}{ylabel}]")
        tx_points = [
            row for row in tx_rows
            if row["tx_power_dbm"] == power and _setting_key(row) == setting
        ]
        rx_points = [row for row in rx_rows if _setting_key(row) == setting]
        lines.extend(
            [
                r"\addplot+[color=blue!75!black, solid, mark=*, mark size=2.2pt,",
                *_energy_coordinates(tx_points),
            ]
        )
        if setting == legend_setting:
            lines.append(rf"\addlegendentry{{TX at {_latex_number(power)} dBm}}")
        lines.extend(
            [
                r"\addplot+[color=red!75!black, dashed, mark=square*, mark size=2.2pt,",
                *_energy_coordinates(rx_points),
            ]
        )
        if setting == legend_setting:
            lines.append(rf"\addlegendentry{{RX, stimulus at {_latex_number(power)} dBm}}")
    note = (
        "Mean of 5 repetitions; error bars show standard deviation. RX energy includes "
        "receiver activation, reception, and processing of all physical frames."
    )
    lines.extend(_energy_document_end(module, "TX--RX comparison", note, len(settings)))
    path.write_text("\n".join(lines), encoding="utf-8")


def _continuous_steps(manifest: dict[str, Any]) -> tuple[Path, Path, list[Path]]:
    tx: Path | None = None
    tx_setting = ""
    rx_by_setting: dict[str, tuple[float, Path]] = {}
    for step in manifest["steps"]:
        if step.get("result_kind") != "continuous":
            continue
        result = Path(step["accepted_result"])
        with (result / "summary.csv").open(encoding="utf-8", newline="") as stream:
            rows = list(csv.DictReader(stream))
        if not rows:
            raise ValueError(f"Empty continuous result: {result}")
        direction = rows[0]["measurement_direction"]
        settings = {_setting_key(row) for row in rows}
        rates = {float(row["bit_rate_kbps"]) for row in rows}
        if len(settings) != 1 or len(rates) != 1:
            raise ValueError(f"Mixed rates in continuous result: {result}")
        setting = settings.pop()
        rate = rates.pop()
        if direction == "tx":
            tx = result
            tx_setting = setting
        elif direction == "rx":
            if setting in rx_by_setting:
                raise ValueError(f"Duplicate continuous RX setting: {setting}")
            rx_by_setting[setting] = (rate, result)
    if tx is None or not rx_by_setting:
        raise ValueError("Missing accepted continuous TX or RX results")
    if tx_setting not in rx_by_setting:
        raise ValueError("No RX sweep matches the continuous TX radio setting")
    ordered = [
        item[1]
        for _setting, item in sorted(
            rx_by_setting.items(), key=lambda pair: (pair[1][0], pair[0])
        )
    ]
    return tx, rx_by_setting[tx_setting][1], ordered


def _read_loss_rows(result_dirs: list[Path]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for result_dir in result_dirs:
        metadata = json.loads((result_dir / "metadata.json").read_text(encoding="utf-8"))
        if metadata.get("measurement_direction") != "rx":
            raise ValueError(f"Not an RX continuous result: {result_dir}")
        with (result_dir / "summary.csv").open(encoding="utf-8", newline="") as stream:
            for raw in csv.DictReader(stream):
                transmitted = int(float(raw["frames_transmitted"]))
                received = int(raw["frames_received"])
                loss = float(raw["frame_loss_percent"])
                row = {
                    "bit_rate_kbps": float(raw["bit_rate_kbps"]),
                    "rf_profile": raw.get("rf_profile", ""),
                    "tx_power_dbm": float(raw["tx_power_dbm"]),
                    "frames_transmitted": transmitted,
                    "frames_received": received,
                    "frames_lost": transmitted - received,
                    "frame_loss_percent": loss,
                    "delivery_percent": 100.0 - loss,
                    "requested_duration_s": float(raw["requested_duration_s"]),
                    "frame_bytes": int(raw["frame_bytes"]),
                    "inter_frame_gap_ms": int(raw["inter_frame_gap_ms"]),
                    "status": raw["status"],
                    "source_directory": result_dir.name,
                }
                expected_loss = (
                    100.0 * (transmitted - received) / transmitted
                    if transmitted
                    else 0.0
                )
                valid_no_rx_data = (
                    row["status"] == "no_rx_data"
                    and transmitted > 0
                    and received == 0
                    and math.isclose(loss, 100.0, abs_tol=1e-6)
                )
                if (
                    row["status"] not in {"ok", "no_rx_data"}
                    or received < 0
                    or received > transmitted
                    or not 0 <= loss <= 100
                    or not math.isclose(loss, expected_loss, abs_tol=1e-4)
                    or (row["status"] == "no_rx_data" and not valid_no_rx_data)
                ):
                    raise ValueError(f"Invalid loss result: {row}")
                rows.append(row)
    rows.sort(
        key=lambda row: (
            row["tx_power_dbm"], row["bit_rate_kbps"], row["rf_profile"]
        )
    )
    powers = {row["tx_power_dbm"] for row in rows}
    settings = {_setting_key(row) for row in rows}
    expected = {(power, setting) for power in powers for setting in settings}
    actual = {(row["tx_power_dbm"], _setting_key(row)) for row in rows}
    if actual != expected or len(rows) != len(expected):
        raise ValueError("Incomplete or duplicate continuous loss matrix")
    return rows


def _style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_loss_reports(base: Path, rows: list[dict[str, Any]], module: str) -> None:
    csv_path = base.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=LOSS_FIELDS)
        writer.writeheader()
        writer.writerows(rows)

    powers = sorted({row["tx_power_dbm"] for row in rows})
    rates = sorted({row["bit_rate_kbps"] for row in rows})
    settings = _radio_settings(rows)
    frame_sizes = sorted({int(row["frame_bytes"]) for row in rows})
    gap_by_rate = {
        rate: sorted(
            {
                int(row["inter_frame_gap_ms"])
                for row in rows
                if row["bit_rate_kbps"] == rate
            }
        )
        for rate in rates
    }
    gap_note = "; ".join(
        f"{_number(rate)} kbps: {'/'.join(map(str, gaps))} ms"
        for rate, gaps in gap_by_rate.items()
    )
    workbook = Workbook()
    results = workbook.active
    results.title = "loss_results"
    results.append(LOSS_FIELDS)
    for row in rows:
        results.append([row[field] for field in LOSS_FIELDS])
    _style_sheet(results)
    matrix = workbook.create_sheet("loss_matrix_percent")
    matrix.append(["TX power [dBm]", *(setting[2] for setting in settings)])
    by_key = {(row["tx_power_dbm"], _setting_key(row)): row for row in rows}
    for power in powers:
        matrix.append(
            [
                power,
                *(by_key[(power, setting[0])]["frame_loss_percent"] for setting in settings),
            ]
        )
    _style_sheet(matrix)
    workbook.save(base.with_suffix(".xlsx"))

    # Logarithmic rates such as 4.8 and 5 kbps are physically almost
    # coincident on the axis.  Put neighbouring labels on alternating rows so
    # both modem profiles remain legible without changing their data position.
    regular_rates: list[float] = []
    staggered_rates: list[float] = []
    previous_rate: float | None = None
    previous_was_staggered = False
    for rate in rates:
        is_close = previous_rate is not None and rate / previous_rate < 1.15
        use_staggered_row = is_close and not previous_was_staggered
        (staggered_rates if use_staggered_row else regular_rates).append(rate)
        previous_rate = rate
        previous_was_staggered = use_staggered_row

    ticks = ",".join(_number(rate) for rate in regular_rates)
    labels = ",".join(_latex_number(rate) for rate in regular_rates)
    staggered_ticks = ",".join(_number(rate) for rate in staggered_rates)
    staggered_labels = ",".join(_latex_number(rate) for rate in staggered_rates)
    maximum = max(row["frame_loss_percent"] for row in rows)
    ymax = min(100, max(10, math.ceil(maximum * 1.25 / 5) * 5))
    xmin = min(rates) / 1.5
    xmax = max(rates) * 1.5
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{axis}[width=12.5cm, height=7.4cm,",
        rf"  title={{{module}: packet loss versus radio rate}},",
        r"  xmode=log, log basis x=10,",
        rf"  xmin={_number(xmin)}, xmax={_number(xmax)}, xtick={{{ticks}}},",
        rf"  xticklabels={{{labels}}}, ymin=0, ymax={_number(ymax)},",
        *(
            [
                rf"  extra x ticks={{{staggered_ticks}}}, extra x tick labels={{{staggered_labels}}},",
                r"  extra x tick style={grid=none, xticklabel style={yshift=-0.35cm}},",
            ]
            if staggered_rates
            else []
        ),
        r"  xlabel={Radio rate [kbps]}, ylabel={Lost frames [\%]},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"  every axis plot/.append style={line width=1.25pt, mark size=2.8pt},",
        r"  legend style={draw=none, fill=white, fill opacity=0.85, text opacity=1,",
        r"    at={(0.03,0.97)}, anchor=north west, legend columns=1},",
        r"]",
    ]
    for index, power in enumerate(reversed(powers)):
        selected = sorted(
            (row for row in rows if row["tx_power_dbm"] == power),
            key=lambda row: (row["bit_rate_kbps"], row["rf_profile"]),
        )
        coordinates = " ".join(
            f"({_number(row['bit_rate_kbps'])},{_number(row['frame_loss_percent'])})"
            for row in selected
        )
        plot_style = "only marks" if len(rates) < len(settings) else LINE_STYLES[index % len(LINE_STYLES)]
        lines.append(
            rf"\addplot+[{COLORS[index % len(COLORS)]}, {plot_style}, mark={MARKERS[index % len(MARKERS)]}] coordinates {{{coordinates}}};"
        )
        lines.append(rf"\addlegendentry{{{_latex_number(power)} dBm}}")
    lines.extend(
        [
            r"\end{axis}",
            r"\node[font=\footnotesize, align=center, text width=12cm, anchor=north]",
            r"  at ([yshift=-1.45cm]current axis.south)",
            rf"  {{60 s per point; {','.join(map(str, frame_sizes))} B frames; inter-frame gap: {gap_note}. Shared rates may contain multiple RF profiles.}};",
            r"\end{tikzpicture}",
            r"\end{document}",
            "",
        ]
    )
    base.with_suffix(".tex").write_text("\n".join(lines), encoding="utf-8")


def _copy_provenance(manifest_path: Path, output_dir: Path) -> None:
    destination = output_dir / "campaign_logs"
    destination.mkdir(parents=True, exist_ok=True)
    session_dir = manifest_path.parent
    for name in (
        "manifest.json",
        "session.log",
        "codex_callback.log",
        "recovery_overrides.json",
    ):
        source = session_dir / name
        if source.exists():
            shutil.copy2(source, destination / name)
    source_logs = session_dir / "logs"
    if source_logs.exists():
        shutil.copytree(source_logs, destination / "attempts", dirs_exist_ok=True)
    recovery_path = session_dir / "recovery_overrides.json"
    if recovery_path.exists():
        recovery = json.loads(recovery_path.read_text(encoding="utf-8"))
        for relative_result in set(recovery.get("steps", {}).values()):
            source = (session_dir / str(relative_result)).resolve()
            if not source.is_relative_to(session_dir.resolve()):
                raise ValueError(f"Recovery result escapes the session: {source}")
            relative = source.relative_to(session_dir.resolve())
            recovery_destination = destination / relative
            recovery_destination.mkdir(parents=True, exist_ok=True)
            for name in ("metadata.json", "summary.csv", "aggregates.csv"):
                item = source / name
                if item.exists():
                    shutil.copy2(item, recovery_destination / name)


def generate(manifest_path: Path, output_dir: Path, base_name: str) -> list[Path]:
    manifest = _read_manifest(manifest_path)
    output_dir.mkdir(parents=True, exist_ok=True)
    tx_rows, tx_summary, tx_metadata = _collect_packet_results(manifest, "tx")
    rx_rows, rx_summary, rx_metadata = _collect_packet_results(manifest, "rx")
    repetitions = int(manifest["config"]["repetitions"])
    sizes, powers, settings, rx_power, module, frame_limit = _validate_energy_matrix(
        tx_rows, rx_rows, repetitions
    )
    base = output_dir / base_name
    tx_base = base.with_name(base.name + "_tx")
    rx_base = base.with_name(base.name + "_rx")
    large_base = base.with_name(base.name + "_payload_128_512_1024")
    tx_campaign_metadata = _campaign_metadata(
        tx_metadata, "tx", sizes, powers, settings
    )
    rx_campaign_metadata = _campaign_metadata(
        rx_metadata, "rx", sizes, [rx_power], settings
    )
    write_transfer_csv(tx_base.with_suffix(".csv"), tx_rows)
    write_transfer_xlsx(
        tx_base.with_suffix(".xlsx"), tx_rows, tx_summary, tx_campaign_metadata
    )
    write_transfer_csv(rx_base.with_suffix(".csv"), rx_rows)
    write_transfer_xlsx(
        rx_base.with_suffix(".xlsx"), rx_rows, rx_summary, rx_campaign_metadata
    )
    large_rows = [row for row in tx_rows if row["payload_bytes"] >= 128]
    large_summary = [row for row in tx_summary if int(row["payload_bytes"]) >= 128]
    if large_rows:
        write_transfer_csv(large_base.with_suffix(".csv"), large_rows)
        write_transfer_xlsx(
            large_base.with_suffix(".xlsx"),
            large_rows,
            large_summary,
            tx_campaign_metadata,
        )
    else:
        large_base.with_suffix(".csv").unlink(missing_ok=True)
        large_base.with_suffix(".xlsx").unlink(missing_ok=True)
    _write_energy_data(
        base.with_name(base.name + "_data").with_suffix(".csv"),
        sorted(
            tx_rows + rx_rows,
            key=lambda row: (
                row["measurement_direction"],
                row["tx_power_dbm"],
                row["bit_rate_kbps"],
                row.get("rf_profile", ""),
                row["payload_bytes"],
            ),
        ),
    )
    _write_energy_data(
        base.with_name(base.name + "_energy_vs_payload_data").with_suffix(".csv"),
        tx_rows,
    )
    _write_tx_tex(
        base.with_name(base.name + "_tx_energy").with_suffix(".tex"),
        tx_rows, sizes, powers, settings, module, frame_limit, "TX energy",
    )
    _write_tx_tex(
        base.with_name(base.name + "_energy_vs_payload").with_suffix(".tex"),
        tx_rows, sizes, powers, settings, module, frame_limit,
        "energy versus logical payload",
    )
    _write_tx_rx_tex(
        base.with_name(base.name + "_tx_rx_energy").with_suffix(".tex"),
        tx_rows, rx_rows, sizes, settings, rx_power, module,
    )

    continuous_tx, continuous_rx, loss_dirs = _continuous_steps(manifest)
    tx_continuous_rows, tx_continuous_metadata = continuous_report.read_session(
        continuous_tx
    )
    rx_continuous_rows, rx_continuous_metadata = continuous_report.read_session(
        continuous_rx
    )
    continuous_report.validate(
        tx_continuous_rows,
        tx_continuous_metadata,
        rx_continuous_rows,
        rx_continuous_metadata,
    )
    all_rx_continuous_rows: list[dict[str, Any]] = []
    for result_dir in loss_dirs:
        rows, _metadata = continuous_report.read_session(result_dir)
        all_rx_continuous_rows.extend(rows)
    continuous_base = base.with_name(base.name + "_continuous")
    continuous_rows = sorted(
        tx_continuous_rows + all_rx_continuous_rows,
        key=lambda row: (
            row["measurement_direction"],
            row.get("rf_profile", ""),
            row["tx_power_dbm"],
        ),
    )
    continuous_report.write_csv(continuous_base.with_suffix(".csv"), continuous_rows)
    continuous_report.write_xlsx(
        continuous_base.with_suffix(".xlsx"),
        tx_continuous_rows,
        rx_continuous_rows,
        tx_continuous_metadata,
        rx_continuous_metadata,
        all_rx_rows=all_rx_continuous_rows,
    )
    continuous_report.write_power_tex(
        continuous_base.with_name(continuous_base.name + "_average_power").with_suffix(".tex"),
        tx_continuous_rows,
        rx_continuous_rows,
        module,
    )
    continuous_report.write_delivery_tex(
        continuous_base.with_name(continuous_base.name + "_rx_delivery").with_suffix(".tex"),
        rx_continuous_rows,
        module,
    )
    loss_rows = _read_loss_rows(loss_dirs)
    _write_loss_reports(
        base.with_name(base.name + "_loss_vs_rate"), loss_rows, module
    )
    _copy_provenance(manifest_path, output_dir)
    return sorted(path for path in output_dir.rglob("*") if path.is_file())


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate compact reports from an unattended web campaign"
    )
    parser.add_argument("manifest", type=Path)
    parser.add_argument("output_dir", type=Path)
    parser.add_argument("--base-name", required=True)
    args = parser.parse_args()
    outputs = generate(args.manifest, args.output_dir, args.base_name)
    for output in outputs:
        print(output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
