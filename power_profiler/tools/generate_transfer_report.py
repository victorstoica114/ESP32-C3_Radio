from __future__ import annotations

import argparse
import csv
import json
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPORT_FIELDS = [
    "profile_id",
    "module",
    "measurement_direction",
    "payload_bytes",
    "frame_count",
    "max_frame_payload_bytes",
    "tx_power_dbm",
    "bit_rate_kbps",
    "rf_profile",
    "runs",
    "events_detected",
    "packets_attempted",
    "packets_received",
    "packets_lost",
    "packet_loss_percent",
    "status_ok_runs",
    "event_duration_ms_mean",
    "event_duration_ms_stdev",
    "tx_mean_uA_mean",
    "tx_peak_uA_mean",
    "rx_mean_uA_mean",
    "rx_peak_uA_mean",
    "event_mean_uA_mean",
    "event_peak_uA_mean",
    "energy_total_uJ_mean",
    "energy_total_uJ_stdev",
    "energy_total_mJ_mean",
    "energy_excess_uJ_mean",
    "energy_per_byte_uJ",
    "energy_excess_per_byte_uJ",
    "effective_payload_rate_kbps",
    "energy_cv_percent",
    "sample_loss_percent_mean",
    "sample_loss_percent_max",
]


def _number(value: Any) -> float:
    return float(value) if value not in (None, "") else 0.0


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as stream:
        return list(csv.DictReader(stream))


def _group_key(row: dict[str, str]) -> tuple[int, int, int, str]:
    return (
        int(row["payload_bytes"]),
        int(row["frame_count"]),
        int(row["max_frame_payload_bytes"]),
        row["parameters_json"],
    )


def _bit_rate_kbps(params: dict[str, Any], metadata: dict[str, Any]) -> float:
    if params.get("bit_rate_kbps") is not None:
        return float(params["bit_rate_kbps"])
    airtime = metadata["profile"].get("airtime", {})
    rate_axis = airtime.get("rate_axis")
    if rate_axis not in params:
        raise ValueError(f"Cannot derive bit rate from parameters: {params}")
    rate_value = params[rate_axis]
    mapping = airtime.get("rate_bps_by_value")
    if mapping is not None:
        try:
            return float(mapping[str(rate_value)]) / 1000.0
        except KeyError as exc:
            raise ValueError(f"No bit-rate mapping for {rate_value!r}") from exc
    return float(rate_value) * float(airtime.get("rate_multiplier", 1000.0)) / 1000.0


def build_report(result_dir: Path) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, Any]]:
    metadata = json.loads((result_dir / "metadata.json").read_text(encoding="utf-8"))
    summary = _read_csv(result_dir / "summary.csv")
    aggregates = _read_csv(result_dir / "aggregates.csv")
    summary_groups: dict[tuple[int, int, int, str], list[dict[str, str]]] = defaultdict(list)
    for row in summary:
        summary_groups[_group_key(row)].append(row)

    report: list[dict[str, Any]] = []
    for aggregate in aggregates:
        params = json.loads(aggregate["parameters_json"])
        direction = aggregate.get("measurement_direction") or metadata.get(
            "measurement_direction", "tx"
        )
        payload_bytes = int(aggregate["payload_bytes"])
        duration_ms = _number(aggregate["event_duration_ms_mean"])
        energy_uJ = _number(aggregate["energy_total_uJ_mean"])
        excess_uJ = _number(aggregate["energy_excess_uJ_mean"])
        energy_stdev = _number(aggregate["energy_total_uJ_stdev"])
        runs = summary_groups[_group_key(aggregate)]
        sample_loss = [_number(row["sample_loss_percent"]) for row in runs]
        report.append(
            {
                "profile_id": aggregate["profile_id"],
                "module": metadata["profile"]["display_name"],
                "measurement_direction": direction,
                "payload_bytes": payload_bytes,
                "frame_count": int(aggregate["frame_count"]),
                "max_frame_payload_bytes": int(aggregate["max_frame_payload_bytes"]),
                "tx_power_dbm": params["tx_power_dbm"],
                "bit_rate_kbps": _bit_rate_kbps(params, metadata),
                "rf_profile": params.get("rf_profile", ""),
                "runs": int(aggregate["runs"]),
                "events_detected": int(aggregate["events_detected"]),
                "packets_attempted": int(aggregate.get("packets_attempted") or 0),
                "packets_received": int(aggregate.get("packets_received") or 0),
                "packets_lost": int(aggregate.get("packets_lost") or 0),
                "packet_loss_percent": _number(
                    aggregate.get("packet_loss_percent")
                ),
                # Missing data at an optional peer does not invalidate a TX
                # current capture. Link loss remains reported separately.
                "status_ok_runs": sum(
                    row["status"] == "ok"
                    or (direction == "tx" and row["status"] == "rx_missing")
                    for row in runs
                ),
                "event_duration_ms_mean": duration_ms,
                "event_duration_ms_stdev": _number(aggregate["event_duration_ms_stdev"]),
                "tx_mean_uA_mean": _number(aggregate["tx_mean_uA_mean"]),
                "tx_peak_uA_mean": _number(aggregate["tx_peak_uA_mean"]),
                "rx_mean_uA_mean": _number(aggregate.get("rx_mean_uA_mean")),
                "rx_peak_uA_mean": _number(aggregate.get("rx_peak_uA_mean")),
                "event_mean_uA_mean": _number(
                    aggregate.get("event_mean_uA_mean", aggregate["tx_mean_uA_mean"])
                ),
                "event_peak_uA_mean": _number(
                    aggregate.get("event_peak_uA_mean", aggregate["tx_peak_uA_mean"])
                ),
                "energy_total_uJ_mean": energy_uJ,
                "energy_total_uJ_stdev": energy_stdev,
                "energy_total_mJ_mean": energy_uJ / 1000.0,
                "energy_excess_uJ_mean": excess_uJ,
                "energy_per_byte_uJ": energy_uJ / payload_bytes,
                "energy_excess_per_byte_uJ": excess_uJ / payload_bytes,
                "effective_payload_rate_kbps": (
                    payload_bytes * 8.0 / duration_ms if duration_ms else 0.0
                ),
                "energy_cv_percent": energy_stdev / energy_uJ * 100.0 if energy_uJ else 0.0,
                "sample_loss_percent_mean": statistics.fmean(sample_loss),
                "sample_loss_percent_max": max(sample_loss),
            }
        )
    report.sort(
        key=lambda row: (
            row["tx_power_dbm"],
            row["bit_rate_kbps"],
            row["rf_profile"],
            row["payload_bytes"],
        )
    )
    return report, summary, metadata


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=REPORT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def _style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        width = min(45, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = width


def _append_rows(sheet, fields: list[str], rows: list[dict[str, Any]]) -> None:
    sheet.append(fields)
    for row in rows:
        sheet.append([row.get(field, "") for field in fields])
    _style_sheet(sheet)


def write_xlsx(
    path: Path,
    report: list[dict[str, Any]],
    summary: list[dict[str, str]],
    metadata: dict[str, Any],
) -> None:
    workbook = Workbook()
    aggregates_sheet = workbook.active
    aggregates_sheet.title = "aggregates"
    _append_rows(aggregates_sheet, REPORT_FIELDS, report)

    matrix_sheet = workbook.create_sheet("energy_matrix_mJ")
    payloads = sorted({int(row["payload_bytes"]) for row in report})
    matrix_fields = ["tx_power_dbm", "bit_rate_kbps"] + [f"{size}_B" for size in payloads]
    matrix_rows = []
    for power, rate in sorted({(row["tx_power_dbm"], row["bit_rate_kbps"]) for row in report}):
        values = {
            int(row["payload_bytes"]): row["energy_total_mJ_mean"]
            for row in report
            if row["tx_power_dbm"] == power and row["bit_rate_kbps"] == rate
        }
        matrix_rows.append(
            {
                "tx_power_dbm": power,
                "bit_rate_kbps": rate,
                **{f"{size}_B": values.get(size, "") for size in payloads},
            }
        )
    _append_rows(matrix_sheet, matrix_fields, matrix_rows)

    summary_sheet = workbook.create_sheet("summary_runs")
    summary_fields = list(summary[0]) if summary else []
    _append_rows(summary_sheet, summary_fields, summary)

    metadata_sheet = workbook.create_sheet("metadata")
    metadata_sheet.append(["field", "value"])
    metadata_sheet.append(["created_utc", metadata.get("created_utc", "")])
    metadata_sheet.append(["module", metadata["profile"]["display_name"]])
    metadata_sheet.append(["profile_id", metadata["profile"]["profile_id"]])
    metadata_sheet.append(
        ["measurement_direction", metadata.get("measurement_direction", "tx")]
    )
    metadata_sheet.append(["measured_port", metadata.get("measured_port", "")])
    metadata_sheet.append(["peer_port", metadata.get("peer_port", "")])
    metadata_sheet.append(["radio_port", metadata.get("radio_port", "")])
    metadata_sheet.append(["ppk_port", metadata.get("ppk_port", "")])
    metadata_sheet.append(["voltage_mv", metadata.get("voltage_mv", "")])
    metadata_sheet.append(["sample_rate_hz", metadata.get("sample_rate_hz", "")])
    frame_limit = metadata["profile"]["transmit"].get("frame_payload_bytes")
    metadata_sheet.append(
        [
            "fragmentation",
            f"Logical transfers use physical frames of at most {frame_limit} B"
            if frame_limit
            else "No explicit fragmentation limit",
        ]
    )
    metadata_sheet.append(
        [
            "receiver_validation",
            "Required" if metadata.get("measurement_direction") == "rx" else "Optional",
        ]
    )
    metadata_sheet.append(["profile_json", json.dumps(metadata["profile"], ensure_ascii=False)])
    _style_sheet(metadata_sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate CSV/XLSX for a profiler result directory")
    parser.add_argument("result_dir", type=Path)
    parser.add_argument("output_base", type=Path)
    parser.add_argument("--sizes", help="optional comma-separated payload sizes")
    args = parser.parse_args()
    report, summary, metadata = build_report(args.result_dir)
    if args.sizes:
        selected_sizes = {int(value.strip()) for value in args.sizes.split(",")}
        report = [row for row in report if row["payload_bytes"] in selected_sizes]
        summary = [
            row for row in summary if int(row["payload_bytes"]) in selected_sizes
        ]
        if not report:
            raise ValueError("The selected payload sizes produced an empty report")
    csv_path = args.output_base.with_suffix(".csv")
    xlsx_path = args.output_base.with_suffix(".xlsx")
    write_csv(csv_path, report)
    write_xlsx(xlsx_path, report, summary, metadata)
    print(csv_path.resolve())
    print(xlsx_path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
