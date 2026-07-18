from __future__ import annotations

import argparse
import csv
import math
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ENERGY_FIELDS = (
    "module",
    "measurement_direction",
    "payload_bytes",
    "tx_power_dbm",
    "bit_rate_kbps",
    "runs",
    "status_ok_runs",
    "packets_received",
    "energy_total_mJ_mean",
    "energy_total_mJ_stdev",
    "event_mean_mA_mean",
    "event_peak_mA_mean",
    "event_duration_ms_mean",
    "sample_loss_percent_max",
)

LOSS_FIELDS = (
    "bit_rate_kbps",
    "tx_power_dbm",
    "frames_transmitted",
    "frames_received",
    "frames_lost",
    "frame_loss_percent",
    "delivery_percent",
    "requested_duration_s",
    "frame_bytes",
    "inter_frame_gap_ms",
    "status",
    "source_directory",
)

COLORS = ("red!75!black", "orange!90!black", "green!50!black")
MARKERS = ("*", "square*", "triangle*")


def _number(value: float | int) -> str:
    return f"{value:.9g}"


def _read_energy_report(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8", newline="") as stream:
        for raw in csv.DictReader(stream):
            rows.append(
                {
                    "module": raw["module"],
                    "measurement_direction": raw["measurement_direction"],
                    "payload_bytes": int(raw["payload_bytes"]),
                    "tx_power_dbm": float(raw["tx_power_dbm"]),
                    "bit_rate_kbps": float(raw["bit_rate_kbps"]),
                    "runs": int(raw["runs"]),
                    "status_ok_runs": int(raw["status_ok_runs"]),
                    "packets_received": int(raw["packets_received"]),
                    "energy_total_mJ_mean": float(raw["energy_total_mJ_mean"]),
                    "energy_total_mJ_stdev": (
                        float(raw["energy_total_uJ_stdev"]) / 1000.0
                    ),
                    "event_mean_mA_mean": (
                        float(raw["event_mean_uA_mean"]) / 1000.0
                    ),
                    "event_peak_mA_mean": (
                        float(raw["event_peak_uA_mean"]) / 1000.0
                    ),
                    "event_duration_ms_mean": float(
                        raw["event_duration_ms_mean"]
                    ),
                    "sample_loss_percent_max": float(
                        raw["sample_loss_percent_max"]
                    ),
                }
            )
    return rows


def _validate_energy(
    tx_rows: list[dict[str, Any]], rx_rows: list[dict[str, Any]]
) -> tuple[list[int], list[float], float, float, str]:
    if not tx_rows or not rx_rows:
        raise ValueError("TX and RX reports must not be empty")
    module = str(tx_rows[0]["module"])
    if any(row["module"] != module for row in tx_rows + rx_rows):
        raise ValueError("TX and RX reports refer to different modules")
    sizes = sorted({int(row["payload_bytes"]) for row in tx_rows})
    powers = sorted({float(row["tx_power_dbm"]) for row in tx_rows})
    rates = sorted({float(row["bit_rate_kbps"]) for row in tx_rows})
    rx_powers = sorted({float(row["tx_power_dbm"]) for row in rx_rows})
    rx_rates = sorted({float(row["bit_rate_kbps"]) for row in rx_rows})
    rx_sizes = sorted({int(row["payload_bytes"]) for row in rx_rows})
    if len(rates) != 1 or rates != rx_rates:
        raise ValueError("E79 reports must use the single supported PHY rate")
    if sizes != rx_sizes or len(rx_powers) != 1 or rx_powers[0] not in powers:
        raise ValueError("RX report must cover all sizes at one TX power")
    tx_expected = {
        (size, power, rates[0])
        for size in sizes
        for power in powers
    }
    tx_actual = {
        (row["payload_bytes"], row["tx_power_dbm"], row["bit_rate_kbps"])
        for row in tx_rows
    }
    rx_expected = {(size, rx_powers[0], rates[0]) for size in sizes}
    rx_actual = {
        (row["payload_bytes"], row["tx_power_dbm"], row["bit_rate_kbps"])
        for row in rx_rows
    }
    if tx_actual != tx_expected or len(tx_rows) != len(tx_expected):
        raise ValueError("Incomplete or duplicate E79 TX energy matrix")
    if rx_actual != rx_expected or len(rx_rows) != len(rx_expected):
        raise ValueError("Incomplete or duplicate E79 RX energy matrix")
    if any(
        row["runs"] != 5 or row["status_ok_runs"] != 5
        for row in tx_rows + rx_rows
    ):
        raise ValueError("Every energy point must contain five valid repetitions")
    if any(row["packets_received"] != 5 for row in rx_rows):
        raise ValueError("Every RX point must contain five received transfers")
    return sizes, powers, rates[0], rx_powers[0], module


def _write_csv(path: Path, fields: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def _energy_limits(rows: list[dict[str, Any]]) -> tuple[float, float]:
    minimum = min(float(row["energy_total_mJ_mean"]) for row in rows)
    maximum = max(float(row["energy_total_mJ_mean"]) for row in rows)
    return (
        10 ** math.floor(math.log10(minimum * 0.75)),
        10 ** math.ceil(math.log10(maximum * 1.25)),
    )


def _coordinates(rows: list[dict[str, Any]]) -> list[str]:
    lines = [r"  error bars/.cd, y dir=both, y explicit,", r"] coordinates {"]
    for row in sorted(rows, key=lambda item: item["payload_bytes"]):
        lines.append(
            f"  ({_number(row['payload_bytes'])},"
            f"{_number(row['energy_total_mJ_mean'])}) +- "
            f"(0,{_number(row['energy_total_mJ_stdev'])})"
        )
    lines.append(r"};")
    return lines


def _group_document_start(
    *, sizes: list[int], columns: int, ymin: float, ymax: float
) -> list[str]:
    ticks = ",".join(str(value) for value in sizes)
    width = 12.0 if columns == 1 else 5.8
    return [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\usepgfplotslibrary{groupplots}",
        r"\usetikzlibrary{calc}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{groupplot}[/pgf/number format/use comma,",
        rf"  group style={{group size={columns} by 1, horizontal sep=1.15cm}},",
        rf"  width={width}cm, height=6.2cm,",
        r"  xmode=log, log basis x={2}, ymode=log, log basis y={10},",
        rf"  ymin={_number(ymin)}, ymax={_number(ymax)},",
        rf"  xmin={_number(sizes[0] * 0.75)}, xmax={_number(sizes[-1] * 1.35)},",
        rf"  xtick={{{ticks}}}, xticklabels={{{ticks}}},",
        r"  x tick label style={rotate=45, anchor=east, font=\small},",
        r"  y tick label style={font=\small},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"  xlabel={Dimensiunea transferului logic [B]},",
        r"  every axis plot/.append style={line width=1.05pt},",
        r"  error bars/error bar style={line width=0.35pt},",
        r"  legend style={font=\small, draw=none, fill=none,",
        r"    at={(0.5,-0.27)}, anchor=north, legend columns=2},",
        r"]",
    ]


def _write_tx_group_tex(
    path: Path,
    tx_rows: list[dict[str, Any]],
    *,
    sizes: list[int],
    powers: list[float],
    rate: float,
    module: str,
    title_suffix: str,
) -> None:
    ymin, ymax = _energy_limits(tx_rows)
    lines = _group_document_start(
        sizes=sizes, columns=len(powers), ymin=ymin, ymax=ymax
    )
    middle = len(powers) // 2
    for index, power in enumerate(powers):
        ylabel = r", ylabel={Energia totală medie [mJ]}" if index == 0 else ""
        lines.append(
            rf"\nextgroupplot[title={{\(P_{{\mathrm{{TX}}}}={_number(power)}\,\mathrm{{dBm}}\)}}{ylabel}]"
        )
        points = [row for row in tx_rows if row["tx_power_dbm"] == power]
        lines.extend(
            [
                r"\addplot+[blue!75!black, solid, mark=*, mark size=2.2pt,",
                *_coordinates(points),
            ]
        )
        if index == middle:
            lines.append(rf"\addlegendentry{{{_number(rate)} kbps}}")
    center = middle + 1
    lines.extend(
        [
            r"\end{groupplot}",
            rf"\node[font=\bfseries\large] at ($(group c{center}r1.north)+(0,0.85cm)$)",
            rf"  {{{module}: {title_suffix}}};",
            r"\node[font=\footnotesize, align=center, text width=18cm]",
            rf"  at ($(group c{center}r1.south)+(0,-3.05cm)$)",
            "  {Media a 5 repetări; barele indică abaterea standard. Transferurile de 128--1024 B\\\\",
            r"   sunt fragmentate în cadre fizice de maximum 64 B. Firmware-ul E79 testat expune o singură viteză PHY: 50 kbps.};",
            r"\end{tikzpicture}",
            r"\end{document}",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_tx_rx_tex(
    path: Path,
    tx_rows: list[dict[str, Any]],
    rx_rows: list[dict[str, Any]],
    *,
    sizes: list[int],
    rate: float,
    rx_power: float,
    module: str,
) -> None:
    tx_selected = [row for row in tx_rows if row["tx_power_dbm"] == rx_power]
    ymin, ymax = _energy_limits(tx_selected + rx_rows)
    lines = _group_document_start(sizes=sizes, columns=1, ymin=ymin, ymax=ymax)
    lines.extend(
        [
            rf"\nextgroupplot[title={{{_number(rate)} kbps}}, ylabel={{Energia totală medie [mJ]}}]",
            r"\addplot+[blue!75!black, solid, mark=*, mark size=2.2pt,",
            *_coordinates(tx_selected),
            rf"\addlegendentry{{TX la {_number(rx_power)} dBm}}",
            r"\addplot+[red!75!black, dashed, mark=square*, mark size=2.2pt,",
            *_coordinates(rx_rows),
            rf"\addlegendentry{{RX, stimul la {_number(rx_power)} dBm}}",
            r"\end{groupplot}",
            r"\node[font=\bfseries\large] at ($(group c1r1.north)+(0,0.85cm)$)",
            rf"  {{{module}: comparație TX--RX}};",
            r"\node[font=\footnotesize, align=center, text width=12cm]",
            r"  at ($(group c1r1.south)+(0,-3.05cm)$)",
            "  {Media a 5 repetări; barele indică abaterea standard. Energia RX include activarea receptorului,\\\\",
            r"   recepția și procesarea tuturor cadrelor transferului logic.};",
            r"\end{tikzpicture}",
            r"\end{document}",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def _read_loss_rows(result_dir: Path) -> tuple[list[dict[str, Any]], str]:
    rows: list[dict[str, Any]] = []
    with (result_dir / "summary.csv").open(
        encoding="utf-8", newline=""
    ) as stream:
        for raw in csv.DictReader(stream):
            transmitted = int(float(raw["frames_transmitted"]))
            received = int(float(raw["frames_received"]))
            loss = float(raw["frame_loss_percent"])
            rows.append(
                {
                    "bit_rate_kbps": float(raw["bit_rate_kbps"]),
                    "tx_power_dbm": float(raw["tx_power_dbm"]),
                    "frames_transmitted": transmitted,
                    "frames_received": received,
                    "frames_lost": transmitted - received,
                    "frame_loss_percent": loss,
                    "delivery_percent": 100.0 - loss,
                    "requested_duration_s": float(raw["requested_duration_s"]),
                    "frame_bytes": int(raw["frame_bytes"]),
                    "inter_frame_gap_ms": int(raw["inter_frame_gap_ms"]),
                    "status": raw["status"],
                    "source_directory": str(result_dir.resolve()),
                }
            )
    rows.sort(key=lambda row: (row["tx_power_dbm"], row["bit_rate_kbps"]))
    if not rows:
        raise ValueError("Continuous RX result is empty")
    rates = {row["bit_rate_kbps"] for row in rows}
    powers = {row["tx_power_dbm"] for row in rows}
    if len(rates) != 1 or len(rows) != len(powers):
        raise ValueError("E79 loss input must contain one row per power at one rate")
    for row in rows:
        if row["requested_duration_s"] != 60.0:
            raise ValueError("Loss measurements must use 60-second windows")
        if row["frame_bytes"] != 32 or row["inter_frame_gap_ms"] != 15:
            raise ValueError("Loss measurements must use 32-byte frames and 15 ms gap")
        if not 0.0 <= row["frame_loss_percent"] <= 100.0:
            raise ValueError(f"Invalid frame loss row: {row}")
    module = "Ebyte E79-400DM2005S, 433.92 MHz"
    return rows, module


def _style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(sheet.columns, start=1):
        width = min(60, max(len(str(cell.value or "")) for cell in column) + 2)
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write_loss_xlsx(path: Path, rows: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    results = workbook.active
    results.title = "loss_results"
    results.append(LOSS_FIELDS)
    for row in rows:
        results.append([row[field] for field in LOSS_FIELDS])
    _style_sheet(results)

    matrix = workbook.create_sheet("loss_matrix_percent")
    rates = sorted({row["bit_rate_kbps"] for row in rows})
    matrix.append(["TX power [dBm]", *rates])
    by_key = {
        (row["tx_power_dbm"], row["bit_rate_kbps"]): row for row in rows
    }
    for power in sorted({row["tx_power_dbm"] for row in rows}):
        matrix.append(
            [
                power,
                *(by_key[(power, rate)]["frame_loss_percent"] for rate in rates),
            ]
        )
    _style_sheet(matrix)
    workbook.save(path)


def _write_loss_tex(
    path: Path, rows: list[dict[str, Any]], module_title: str
) -> None:
    rate = float(rows[0]["bit_rate_kbps"])
    powers = sorted({float(row["tx_power_dbm"]) for row in rows})
    lines = [
        r"\documentclass[tikz,border=6pt]{standalone}",
        r"\usepackage[T1]{fontenc}",
        r"\usepackage[utf8]{inputenc}",
        r"\usepackage{pgfplots}",
        r"\pgfplotsset{compat=1.18}",
        r"\begin{document}",
        r"\begin{tikzpicture}",
        r"\begin{axis}[width=12.5cm, height=7.4cm,",
        rf"  title={{{module_title}: packet loss în funcție de viteza radio}},",
        rf"  xmin={_number(rate * 0.8)}, xmax={_number(rate * 1.2)},",
        rf"  xtick={{{_number(rate)}}}, xticklabels={{{_number(rate)}}},",
        r"  ymin=0, ymax=105, ytick={0,20,40,60,80,100},",
        r"  xlabel={Viteza radio [kbps]}, ylabel={Cadre pierdute [\%]},",
        r"  grid=both, minor grid style={gray!15}, major grid style={gray!35},",
        r"  every axis plot/.append style={only marks, mark size=3.2pt},",
        r"  legend style={draw=none, fill=white, fill opacity=0.85, text opacity=1,",
        r"    at={(0.03,0.97)}, anchor=north west, legend columns=1},",
        r"]",
    ]
    for index, power in enumerate(reversed(powers)):
        row = next(item for item in rows if item["tx_power_dbm"] == power)
        color = COLORS[(len(powers) - 1 - index) % len(COLORS)]
        marker = MARKERS[(len(powers) - 1 - index) % len(MARKERS)]
        lines.append(
            rf"\addplot+[{color}, mark={marker}] coordinates "
            rf"{{({_number(rate)},{_number(row['frame_loss_percent'])})}};"
        )
        lines.append(rf"\addlegendentry{{{_number(power)} dBm}}")
    lines.extend(
        [
            r"\end{axis}",
            r"\node[font=\footnotesize, align=center, text width=12cm, anchor=north]",
            r"  at ([yshift=-1.45cm]current axis.south)",
            "  {60 s/punct; cadre de 32 B; pauză de 15 ms. Firmware-ul E79 testat expune doar 50 kbps,\\\\",
            r"   de aceea graficul conține un singur punct de viteză pentru fiecare putere.};",
            r"\end{tikzpicture}",
            r"\end{document}",
            "",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate E79 energy and single-rate loss comparison reports"
    )
    parser.add_argument("tx_report", type=Path)
    parser.add_argument("rx_report", type=Path)
    parser.add_argument("continuous_rx_result", type=Path)
    parser.add_argument("output_base", type=Path)
    args = parser.parse_args()

    tx_rows = _read_energy_report(args.tx_report)
    rx_rows = _read_energy_report(args.rx_report)
    sizes, powers, rate, rx_power, module = _validate_energy(tx_rows, rx_rows)
    args.output_base.parent.mkdir(parents=True, exist_ok=True)

    combined = sorted(
        tx_rows + rx_rows,
        key=lambda row: (
            row["measurement_direction"],
            row["tx_power_dbm"],
            row["payload_bytes"],
        ),
    )
    data_path = args.output_base.with_name(
        args.output_base.name + "_data"
    ).with_suffix(".csv")
    energy_data_path = args.output_base.with_name(
        args.output_base.name + "_energy_vs_payload_data"
    ).with_suffix(".csv")
    _write_csv(data_path, ENERGY_FIELDS, combined)
    _write_csv(energy_data_path, ENERGY_FIELDS, tx_rows)

    tx_tex = args.output_base.with_name(
        args.output_base.name + "_tx_energy"
    ).with_suffix(".tex")
    tx_rx_tex = args.output_base.with_name(
        args.output_base.name + "_tx_rx_energy"
    ).with_suffix(".tex")
    energy_tex = args.output_base.with_name(
        args.output_base.name + "_energy_vs_payload"
    ).with_suffix(".tex")
    _write_tx_group_tex(
        tx_tex,
        tx_rows,
        sizes=sizes,
        powers=powers,
        rate=rate,
        module=module,
        title_suffix="consumul în transmisie",
    )
    _write_tx_rx_tex(
        tx_rx_tex,
        tx_rows,
        rx_rows,
        sizes=sizes,
        rate=rate,
        rx_power=rx_power,
        module=module,
    )
    _write_tx_group_tex(
        energy_tex,
        tx_rows,
        sizes=sizes,
        powers=powers,
        rate=rate,
        module=module,
        title_suffix="energia în funcție de dimensiune",
    )

    loss_rows, loss_module = _read_loss_rows(args.continuous_rx_result)
    loss_base = args.output_base.with_name(
        args.output_base.name + "_loss_vs_rate"
    )
    loss_csv = loss_base.with_suffix(".csv")
    loss_xlsx = loss_base.with_suffix(".xlsx")
    loss_tex = loss_base.with_suffix(".tex")
    _write_csv(loss_csv, LOSS_FIELDS, loss_rows)
    _write_loss_xlsx(loss_xlsx, loss_rows)
    _write_loss_tex(loss_tex, loss_rows, loss_module)

    for output in (
        data_path,
        energy_data_path,
        tx_tex,
        tx_rx_tex,
        energy_tex,
        loss_csv,
        loss_xlsx,
        loss_tex,
    ):
        print(output.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
