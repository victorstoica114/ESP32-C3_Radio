from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


TOOLS_DIR = Path(__file__).resolve().parents[1] / "tools"
if str(TOOLS_DIR) not in sys.path:
    sys.path.insert(0, str(TOOLS_DIR))

import generate_continuous_report as continuous_report  # noqa: E402
import generate_transfer_report as transfer_report  # noqa: E402
import generate_lora_campaign_reports as lora_reports  # noqa: E402
import generate_lora_variant_comparison as variant_comparison  # noqa: E402
import render_lora_campaign_plots as lora_renderer  # noqa: E402
import render_web_campaign_plots as web_renderer  # noqa: E402
import generate_web_campaign_reports as campaign_reports  # noqa: E402


def _continuous_row(profile: str, power: float) -> dict[str, object]:
    return {
        "measurement_direction": "rx",
        "rf_profile": profile,
        "tx_power_dbm": power,
        "mean_current_uA": 7000.0,
        "mean_power_mW": 23.1,
        "mean_excess_power_mW": 0.0,
        "energy_60s_mJ": 1386.0,
        "frames_received": 100,
        "frames_transmitted": 100.0,
        "frame_loss_percent": 0.0,
    }


class ReportTests(unittest.TestCase):
    def test_generic_loss_report_accepts_verified_total_radio_loss(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            result = Path(temporary)
            (result / "metadata.json").write_text(
                '{"measurement_direction":"rx"}\n', encoding="utf-8"
            )
            (result / "summary.csv").write_text(
                "frames_transmitted,frames_received,frame_loss_percent,"
                "bit_rate_kbps,rf_profile,tx_power_dbm,requested_duration_s,"
                "frame_bytes,inter_frame_gap_ms,status\n"
                "3366,0,100,250,,0,60,32,15,no_rx_data\n",
                encoding="utf-8",
            )

            rows = campaign_reports._read_loss_rows([result])

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["frames_lost"], 3366)
        self.assertEqual(rows[0]["delivery_percent"], 0.0)

    def test_generic_energy_matrix_keeps_valid_current_captures_with_rf_loss(self) -> None:
        common = {
            "module": "nRF24L01",
            "payload_bytes": 32,
            "tx_power_dbm": 0.0,
            "bit_rate_kbps": 250.0,
            "rf_profile": "",
            "runs": 5,
            "status_ok_runs": 5,
            "max_frame_payload_bytes": 32,
        }
        tx_rows = [{**common, "packets_received": 2}]
        rx_rows = [{**common, "packets_received": 4}]

        sizes, powers, settings, rx_power, module, frame_limit = (
            campaign_reports._validate_energy_matrix(tx_rows, rx_rows, 5)
        )

        self.assertEqual(sizes, [32])
        self.assertEqual(powers, [0.0])
        self.assertEqual(settings[0][1], 250.0)
        self.assertEqual(rx_power, 0.0)
        self.assertEqual(module, "nRF24L01")
        self.assertEqual(frame_limit, 32)

    def test_lora_variant_comparison_reports_relative_changes(self) -> None:
        continuous_classic = [
            {
                "measurement_direction": "tx",
                "tx_power_dbm": "10",
                "spreading_factor": "9",
                "bandwidth_khz": "125",
                "mean_current_uA": "20000",
                "mean_power_mW": "66",
                "frames_transmitted": "100",
                "frames_received": "",
            }
        ]
        continuous_variant = [dict(continuous_classic[0])]
        continuous_variant[0]["mean_current_uA"] = "19000"
        continuous_variant[0]["mean_power_mW"] = "62.7"
        packet_classic = [{
            "measurement_direction": "tx",
            "payload_bytes": "8",
            "tx_power_dbm": "10",
            "spreading_factor": "9",
            "bandwidth_khz": "125",
            "energy_total_mJ_mean": "10",
            "energy_cv_percent": "0.2",
            "packets_received": "5",
        }]
        packet_variant = [dict(packet_classic[0])]
        packet_variant[0]["energy_total_mJ_mean"] = "9.5"

        continuous = variant_comparison.compare_continuous(
            continuous_classic, continuous_variant
        )
        packets = variant_comparison.compare_packets(packet_classic, packet_variant)

        self.assertAlmostEqual(continuous[0]["current_delta_percent"], -5.0)
        self.assertAlmostEqual(continuous[0]["power_delta_percent"], -5.0)
        self.assertAlmostEqual(packets[0]["energy_delta_percent"], -5.0)

    def test_lora_loss_matrix_accepts_positive_minimum_tx_power(self) -> None:
        continuous = []
        for power in (2.0, 10.0, 20.0):
            for spreading_factor, transmitted in ((7, 427), (9, 139), (12, 20)):
                continuous.append(
                    {
                        "measurement_direction": "rx",
                        "frames_transmitted": transmitted,
                        "frames_received": transmitted,
                        "requested_duration_s": 60.0,
                        "frame_bytes": 64,
                        "frame_loss_percent": 0.0,
                        "spreading_factor": spreading_factor,
                        "bandwidth_khz": 125.0,
                        "tx_power_dbm": power,
                        "inter_frame_gap_ms": 15,
                        "status": "ok",
                        "source_directory": "recovery",
                    }
                )

        loss = lora_reports._loss_rows(continuous)

        self.assertEqual(len(loss), 9)
        self.assertEqual({row["tx_power_dbm"] for row in loss}, {2.0, 10.0, 20.0})
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "loss.tex"
            lora_reports._write_loss_tex(path, loss, "RA-01H")
            latex = path.read_text(encoding="utf-8")
            self.assertIn(r"\addlegendentry{2 dBm}", latex)

    def test_dependency_free_pdf_renderer_writes_a_valid_document(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "plot.pdf"
            canvas = lora_renderer.PdfCanvas(320, 180)
            canvas.text(20, 140, "Test plot", bold=True)
            canvas.line(20, 20, 300, 150)
            canvas.save(path)

            data = path.read_bytes()
            self.assertTrue(data.startswith(b"%PDF-1.4"))
            self.assertTrue(data.endswith(b"%%EOF\n"))
            self.assertTrue(path.with_suffix(".png").read_bytes().startswith(b"\x89PNG"))

    def test_generic_radio_energy_renderer_writes_pdf_and_png(self) -> None:
        rows = []
        for power in (-1.0, 8.0, 20.0):
            for rate in (0.5, 15.0, 250.0):
                for payload in (8.0, 32.0, 60.0):
                    rows.append(
                        {
                            "tx_power_dbm": power,
                            "bit_rate_kbps": rate,
                            "payload_bytes": payload,
                            "energy_total_mJ_mean": 1.0 + payload / rate,
                        }
                    )
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            path = web_renderer._energy_plot(
                output,
                "test",
                "Generic radio",
                rows,
                title="TX energy",
                suffix="tx_energy",
            )

            self.assertTrue(path.read_bytes().startswith(b"%PDF-1.4"))
            self.assertTrue(path.with_suffix(".png").read_bytes().startswith(b"\x89PNG"))

    def test_energy_plot_scale_includes_sub_millijoule_measurements(self) -> None:
        limits, ticks = lora_renderer._log_energy_scale([0.04, 88.0])

        self.assertEqual(limits, (0.01, 1000.0))
        self.assertEqual(ticks, (0.01, 0.1, 1.0, 10.0, 100.0, 1000.0))

    def test_continuous_plot_fits_a_long_module_title(self) -> None:
        rows = []
        for direction in ("tx", "rx"):
            for power in (-4.0, 10.0, 20.0):
                rows.append(
                    {
                        "measurement_direction": direction,
                        "spreading_factor": "9",
                        "tx_power_dbm": power,
                        "mean_power_mW": 40.0,
                        "mean_current_uA": 12000.0,
                    }
                )
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary)
            path = lora_renderer._continuous(
                output,
                "test",
                "A very long LoRa module variant display name with qualifiers",
                rows,
            )

            self.assertTrue(path.is_file())
            self.assertTrue(path.with_suffix(".png").is_file())

    def test_lora_report_accepts_verified_total_radio_loss(self) -> None:
        row = {
            "status": "no_rx_data",
            "measurement_direction": "rx",
            "frames_transmitted": 450.0,
            "frames_received": 0,
            "frame_loss_percent": 100.0,
            "transmitter_response": "TXCONT=60000,FRAMES=450 | SERIAL_ERRORS=0",
        }

        self.assertTrue(lora_reports._valid_continuous_status(row))

        row["transmitter_response"] = "TXCONT=60000,FRAMES=450 | SERIAL_ERRORS=1"
        self.assertFalse(lora_reports._valid_continuous_status(row))

    def test_lora_manifest_accepts_valid_targeted_recovery(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            session = Path(temporary)
            result = session / "recovery" / "continuous_tx"
            result.mkdir(parents=True)
            (result / "summary.csv").write_text("status\nok\n", encoding="utf-8")
            (result / "metadata.json").write_text("{}\n", encoding="utf-8")
            manifest_path = session / "manifest.json"
            manifest_path.write_text(
                '{"kind":"campaign","state":"completed_with_errors",'
                '"completed_steps":0,"failed_steps":1,"steps":['
                '{"step_id":"continuous_tx_average","status":"failed"}]}',
                encoding="utf-8",
            )
            (session / "recovery_overrides.json").write_text(
                '{"reason":"targeted retry","steps":{'
                '"continuous_tx_average":"recovery/continuous_tx"}}',
                encoding="utf-8",
            )

            manifest = lora_reports._manifest(manifest_path)

            self.assertEqual(manifest["state"], "completed")
            self.assertEqual(manifest["completed_steps"], 1)
            self.assertEqual(manifest["failed_steps"], 0)
            self.assertEqual(
                Path(manifest["steps"][0]["accepted_result"]), result.resolve()
            )
            self.assertEqual(
                manifest["recovery_overrides"]["reason"], "targeted retry"
            )

    def test_generic_campaign_manifest_accepts_valid_targeted_recovery(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            session = Path(temporary)
            result = session / "recovery" / "continuous_tx"
            result.mkdir(parents=True)
            (result / "summary.csv").write_text("status\nok\n", encoding="utf-8")
            (result / "metadata.json").write_text("{}\n", encoding="utf-8")
            manifest_path = session / "manifest.json"
            manifest_path.write_text(
                '{"kind":"campaign","state":"stopped",'
                '"completed_steps":0,"failed_steps":0,"steps":['
                '{"step_id":"continuous_tx_average","status":"stopped"}]}',
                encoding="utf-8",
            )
            (session / "recovery_overrides.json").write_text(
                '{"reason":"targeted retry","steps":{'
                '"continuous_tx_average":"recovery/continuous_tx"}}',
                encoding="utf-8",
            )

            manifest = campaign_reports._read_manifest(manifest_path)

            self.assertEqual(manifest["state"], "completed")
            self.assertEqual(manifest["completed_steps"], 1)
            self.assertEqual(manifest["failed_steps"], 0)
            self.assertEqual(
                Path(manifest["steps"][0]["accepted_result"]), result.resolve()
            )
            self.assertEqual(
                manifest["recovery_overrides"]["reason"], "targeted retry"
            )

    def test_generic_campaign_provenance_includes_recovery_summaries(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            session = Path(temporary) / "session"
            result = session / "recovery" / "continuous_tx"
            result.mkdir(parents=True)
            (result / "summary.csv").write_text("status\nok\n", encoding="utf-8")
            (result / "metadata.json").write_text("{}\n", encoding="utf-8")
            manifest_path = session / "manifest.json"
            manifest_path.write_text("{}\n", encoding="utf-8")
            (session / "recovery_overrides.json").write_text(
                '{"steps":{"continuous_tx_average":"recovery/continuous_tx"}}',
                encoding="utf-8",
            )
            output = Path(temporary) / "output"

            campaign_reports._copy_provenance(manifest_path, output)

            copied = output / "campaign_logs" / "recovery" / "continuous_tx"
            self.assertTrue((copied / "summary.csv").is_file())
            self.assertTrue((copied / "metadata.json").is_file())

    def test_continuous_workbook_exports_all_rx_profiles(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "continuous.xlsx"
            tx = _continuous_row("SLR2K5", -20.0)
            tx["measurement_direction"] = "tx"
            matching_rx = _continuous_row("SLR2K5", -20.0)
            other_rx = _continuous_row("GFSK200", -20.0)

            continuous_report.write_xlsx(
                path,
                [tx],
                [matching_rx],
                {},
                {},
                all_rx_rows=[matching_rx, other_rx],
            )

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(workbook["continuous_results"].max_row - 1, 3)
                self.assertEqual(workbook["comparison"].max_row - 1, 1)
            finally:
                workbook.close()

    def test_transfer_workbook_sanitizes_xml_control_characters(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "transfer.xlsx"
            report = [
                {
                    "payload_bytes": 8,
                    "tx_power_dbm": 20.0,
                    "bit_rate_kbps": 0.3,
                    "energy_total_mJ_mean": 1.0,
                }
            ]
            summary = [{"receiver_response": "\x00012345", "status": "ok"}]
            metadata = {
                "profile": {
                    "display_name": "E32",
                    "profile_id": "RADIO_EBYTE",
                    "transmit": {"frame_payload_bytes": 58},
                }
            }

            transfer_report.write_xlsx(path, report, summary, metadata)

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                self.assertEqual(
                    workbook["summary_runs"]["A2"].value,
                    "012345",
                )
            finally:
                workbook.close()

    def test_loss_graph_staggers_nearly_identical_rate_labels(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            base = Path(temporary) / "loss_vs_rate"
            rows = []
            for profile, rate in (("GFSK4K8", 4.8), ("SLR5", 5.0)):
                rows.append(
                    {
                        "bit_rate_kbps": rate,
                        "rf_profile": profile,
                        "tx_power_dbm": -20.0,
                        "frames_transmitted": 100,
                        "frames_received": 100,
                        "frames_lost": 0,
                        "frame_loss_percent": 0.0,
                        "delivery_percent": 100.0,
                        "requested_duration_s": 60.0,
                        "frame_bytes": 64,
                        "inter_frame_gap_ms": 15,
                        "status": "ok",
                        "source_directory": profile,
                    }
                )

            campaign_reports._write_loss_reports(base, rows, "Test module")

            latex = base.with_suffix(".tex").read_text(encoding="utf-8")
            self.assertIn("xtick={4.8}", latex)
            self.assertIn("extra x ticks={5}", latex)
            self.assertIn("(4.8,0)", latex)
            self.assertIn("(5,0)", latex)


if __name__ == "__main__":
    unittest.main()
